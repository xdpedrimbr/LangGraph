from __future__ import annotations

import os
import random
import re
import math
from difflib import SequenceMatcher
from datetime import date, datetime, timedelta
from typing import Annotated, Optional

from dotenv import load_dotenv
from langgraph.graph import END, START, StateGraph
from langgraph.graph.message import add_messages
from typing_extensions import TypedDict

from solutions.sql_analytics.catalog_loader import load_schema_catalog
from shared.db_client import cleanup_sql, get_cnpj_for_phone, parse_cols, parse_rows, run_query
from solutions.sql_analytics.schema_tools import validate_sql_against_schema
from solutions.sql_analytics.sql_generator import ExtractedParams, SqlGenerator

load_dotenv()

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
SCHEMA_CATALOG_PATH = (os.getenv("SCHEMA_CATALOG_PATH") or os.path.join(_THIS_DIR, "schema_catalog.json")).strip()

_schema = load_schema_catalog(SCHEMA_CATALOG_PATH)
_generator = SqlGenerator(schema=_schema)


# ── Estado do grafo ────────────────────────────────────────────────────────────

class State(TypedDict, total=False):
    messages:   Annotated[list, add_messages]
    phone:      str
    cnpj:       Optional[str]
    question:   str

    # curto-circuito (sem query)
    direct_reply: Optional[str]
    intent:       Optional[str]

    # extração de parâmetros
    extracted_params: Optional[dict]

    # geração de SQL
    sql:              Optional[str]
    sql_attempts:     int
    sql_error:        Optional[str]          # feedback de erro para retry
    supervisor_retry: bool                   # supervisor pediu retry

    # resultado da query
    columns: list
    rows:    list

    # nota do sql gen (ex: "exibindo só top 10")
    sql_note: Optional[str]

    # notas de qualidade de dados (supervisor → writer, sem forçar retry)
    data_quality_notes: Optional[list]

    # contexto do turno anterior (para resolver follow-ups)
    last_question:         Optional[str]
    last_answer:           Optional[str]
    last_extracted_params: Optional[dict]
    is_followup:           bool

    # saída final
    answer: Optional[str]
    output_messages: Optional[list[str]]
    output_messages_question: Optional[str]

    # URL de arquivo para enviar como documento (ex: Excel de perdas)
    excel_url: Optional[str]

    # insight gerado após a resposta
    insight: Optional[str]

    # memória de longo prazo (perfil do lojista)
    profile_hint: Optional[str]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _last_user_message(messages: list) -> str:
    for m in reversed(messages or []):
        if isinstance(m, dict):
            if (m.get("role") or "").lower() in ("user", "human"):
                return str(m.get("content") or "").strip()
        else:
            role = (getattr(m, "type", None) or getattr(m, "role", None) or "").lower()
            if role in ("human", "user"):
                return str(getattr(m, "content", "") or "").strip()
    return ""



# ── Log helper ─────────────────────────────────────────────────────────────────

def _log(node: str, **kv):
    items = " | ".join(f"{k}={v}" for k, v in kv.items())
    print(f"[{node}] {items}")


_VALIDADE_KEYWORDS = [
    "vencimento", "vencendo", "vencer", "validade", "prazo de validade",
    "data de validade", "proximo ao vencimento", "proximos ao vencimento",
    "produto vencido", "produtos vencidos", "vencidos", "vencido",
    "proximo de vencer", "proximos de vencer", "quase vencendo",
]


def _is_validade_question(message: str) -> bool:
    import unicodedata
    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )
    return any(kw in msg_norm for kw in _VALIDADE_KEYWORDS)


_TUDO_BEM_KEYWORDS = [
    "tudo bem", "tudo certo", "tudo bom", "td bem", "td bom", "td certo",
    "como vai", "como esta", "como você esta", "como voce esta",
    "tá funcionando", "ta funcionando", "está funcionando", "esta funcionando",
    "funcionando bem", "tudo por ai", "tudo por aí", "tudo na paz",
]


def _is_tudo_bem_question(message: str) -> bool:
    """Detecta saudações do tipo "tá tudo bem com você?" / "tá funcionando bem?"."""
    import unicodedata
    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )
    return any(kw in msg_norm for kw in _TUDO_BEM_KEYWORDS)


_FAREWELL_KEYWORDS = [
    "tchau", "obrigado", "obrigada", "muito obrigado", "muito obrigada",
    "valeu", "ate mais", "até mais", "ate logo", "até logo", "falou",
    "ate a proxima", "até a próxima", "ate amanha", "até amanhã", "boa noite e tchau",
]


def _is_farewell_question(message: str) -> bool:
    """Detecta despedidas/agradecimentos finais ("tchau", "muito obrigado")."""
    import unicodedata
    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )
    return any(kw in msg_norm for kw in _FAREWELL_KEYWORDS)


_CAPACIDADES_RESUMO = (
    "faturamento, ticket médio e transações, produtos mais/menos vendidos, "
    "estoque, comparação com o mercado, previsões e muito mais"
)


# ── Nós ────────────────────────────────────────────────────────────────────────

async def preprocess_node(state: State) -> dict:
    """Resolve CNPJ e classifica a mensagem."""
    phone    = state.get("phone") or ""
    messages = state.get("messages") or []
    question = _last_user_message(messages)

    _log("preprocess", phone=phone, question=question[:80])

    # Resolve CNPJ pelo telefone
    cnpj = state.get("cnpj") or ""
    if not cnpj:
        try:
            cnpj = await get_cnpj_for_phone(phone)
        except Exception:
            _log("preprocess", result="ERRO ao buscar CNPJ")
            return {
                "question":     question,
                "cnpj":         "",
                "direct_reply": "Não consegui identificar seu estabelecimento. Tente novamente em instantes.",
            }

    if not cnpj:
        _log("preprocess", result="CNPJ não encontrado")
        return {
            "question":     question,
            "cnpj":         "",
            "direct_reply": "Seu número não está vinculado a nenhum estabelecimento. Entre em contato com o suporte.",
        }

    # Verifica se usuário está respondendo ao pedido de tema de insight
    from shared.session_store import get_pending_insight_theme
    if get_pending_insight_theme(phone):
        _log("preprocess", result="pending_insight_theme → insight_theme")
        return {
            "question": question,
            "cnpj":     cnpj,
            "intent":   "insight_theme",
            "is_followup": False,
            "direct_reply": None,
        }

    # Carrega perfil do lojista para personalização
    from shared.profile_store import get_profile, build_profile_hint
    _profile      = get_profile(cnpj) if cnpj else {}
    _profile_hint = build_profile_hint(_profile)

    # Detecta follow-up ANTES de qualquer escrita (last_question reflete turno anterior)
    last_q      = state.get("last_question")
    is_followup = bool(last_q)
    last_answer = state.get("last_answer") or ""

    normalized_question = _normalize_pt(question)
    if normalized_question.startswith("resum"):
        _log("preprocess", result="intent=reposicao_resumo_hortifruti")
        return {
            "question": question,
            "cnpj": cnpj,
            "intent": "reposicao_resumo_hortifruti",
            "is_followup": is_followup,
            "direct_reply": None,
        }

    # Perguntas sobre validade/vencimento de produto — sem dados disponíveis
    if _is_validade_question(question):
        _user_name_v: str | None = None
        try:
            from shared.db_client import get_user_name_for_phone
            _user_name_v = await get_user_name_for_phone(phone, cnpj)
        except Exception:
            pass
        _greeting = f"*{_user_name_v}*, ainda" if _user_name_v else "Ainda"
        _log("preprocess", result="validade/vencimento → sem dados")
        return {
            "question":    question,
            "cnpj":        cnpj,
            "intent":      "out_of_scope",
            "is_followup": is_followup,
            "direct_reply": (
                f"{_greeting} não possuo dados sobre as validades dos seus produtos. "
                "Não consigo te responder sobre esse assunto."
            ),
        }

    # Despedidas/agradecimentos finais — resposta fixa sem chamar o LLM
    # (limitado a mensagens curtas para não disparar em "obrigado, mas me mostre...")
    if len(_words_pt(question)) <= 6 and _is_farewell_question(question):
        _log("preprocess", result="farewell → despedida")
        return {
            "question":     question,
            "cnpj":         cnpj,
            "intent":       "out_of_scope",
            "is_followup":  is_followup,
            "direct_reply": (
                "Muito obrigado pelo seu contato! 😊 Volte sempre que precisar tirar "
                "suas dúvidas ou consultar os dados do seu negócio. Até a próxima!"
            ),
        }

    # Sugestão de promoção — fast path por palavra-chave (preço/ofertas)
    if _is_promocao_question(question):
        _log("preprocess", result="intent=promocao (keyword trigger)")
        return {"question": question, "cnpj": cnpj, "intent": "promocao", "is_followup": is_followup, "direct_reply": None}

    # Relatório de perdas/histórico de movimentações manuais
    if _is_perdas_question(question):
        _log("preprocess", result="intent=perdas_historico (keyword trigger)")
        return {"question": question, "cnpj": cnpj, "intent": "perdas_historico", "is_followup": is_followup, "direct_reply": None}

    # Comandos de estoque por seção — detectados por palavra-chave
    # Fast path: trigger exato por palavra-chave (evita chamar o LLM)
    _cmd = _detect_estoque_command(question)
    if _cmd:
        mode, section = _cmd
        intent = f"{mode}_{section}"
        _log("preprocess", result=f"intent={intent} (keyword trigger)")
        return {"question": question, "cnpj": cnpj, "intent": intent, "is_followup": is_followup, "direct_reply": None}

    # Pedidos como "reposição da banana" não citam necessariamente "hortifruti",
    # mas devem usar o mesmo cálculo e formato de sugestão de compra por item.
    if _is_reposicao_produto_especifico(question):
        _log("preprocess", result="intent=reposicao_hortifruti (produto específico)")
        return {
            "question": question,
            "cnpj": cnpj,
            "intent": "reposicao_hortifruti",
            "is_followup": is_followup,
            "direct_reply": None,
        }

    classify = await _generator.classify(question, last_question=last_q)
    _log("preprocess", cnpj=cnpj, intent=classify.intent)

    # Intents de estoque retornados pelo classificador (follow-ups, variações de frase)
    _ESTOQUE_INTENTS = ("relatorio_hortifruti", "reposicao_hortifruti", "reposicao_mais_hortifruti")
    if classify.intent in _ESTOQUE_INTENTS:
        _log("preprocess", result=f"intent={classify.intent} (classifier)")
        return {"question": question, "cnpj": cnpj, "intent": classify.intent, "is_followup": is_followup, "direct_reply": None}

    if classify.intent == "promocao":
        _log("preprocess", result="intent=promocao (classifier)")
        return {"question": question, "cnpj": cnpj, "intent": "promocao", "is_followup": is_followup, "direct_reply": None}

    if classify.intent != "data_query":
        direct = classify.direct_reply or "Como posso te ajudar?"

        # Personaliza saudação com o nome do usuário (se disponível na tabela de telefones)
        if classify.intent == "greeting" and phone:
            try:
                from shared.db_client import get_user_name_for_phone
                user_name = await get_user_name_for_phone(phone, cnpj)
            except Exception:
                user_name = None

            if _is_tudo_bem_question(question):
                nome_part = f", *{user_name}*" if user_name else ""
                direct = (
                    f"Está tudo bem{nome_part}, e com você? "
                    f"O que gostaria de me perguntar hoje? "
                    f"Posso te responder sobre {_CAPACIDADES_RESUMO}. "
                    f"Por exemplo: _\"Quanto vendi essa semana?\"_"
                )
            elif user_name:
                direct = direct.replace("Olá! ", f"Olá, *{user_name}*! ", 1)

        return {
            "question":     question,
            "cnpj":         cnpj,
            "intent":       classify.intent,
            "is_followup":  is_followup,
            "direct_reply": direct,
        }

    # Reseta intent e direct_reply para data_query (evita vazamento do estado anterior)
    return {
        "question":         question,
        "cnpj":             cnpj,
        "intent":           "data_query",
        "is_followup":      is_followup,
        "direct_reply":     None,
        "extracted_params": None,
        "sql":              None,
        "sql_attempts":     0,
        "sql_error":        None,
        "supervisor_retry": False,
        "columns":          [],
        "rows":             [],
        "profile_hint":     _profile_hint,
    }


# ── Rotação de temas para pedidos vagos de "insight" ──────────────────────────
# Pedidos como "me dê um insight" (sem métrica/produto/categoria/período
# específicos) sempre caíam em faturamento dos últimos 30 dias — repetitivo.
# Aqui rotacionamos entre alguns temas úteis, evitando repetir o último usado
# pelo mesmo número.
_VAGUE_INSIGHT_TEMAS: list[tuple[str, str]] = [
    ("faturamento",          "uma visão geral do faturamento e das transações recentes"),
    ("ticket_medio",         "como está o ticket médio recente"),
    ("diagnostico_positivo", "o que está indo bem na loja"),
    ("diagnostico_negativo", "o que está indo mal e precisa de atenção"),
    ("top_categorias",       "quais categorias mais vendem"),
]

_VAGUE_INSIGHT_GATILHOS = {
    "insight", "analise", "análise", "resumo", "panorama", "diagnostico", "diagnóstico", "dica", "visao", "visão",
}


def _eh_pedido_vago_insight(question: str) -> bool:
    """Detecta pedidos genéricos de "insight"/"análise"/"resumo" — sem métrica,
    produto, categoria ou período específicos — usados para variar o tema do
    insight entregue (em vez de sempre cair em faturamento dos últimos 30 dias)."""
    palavras = _words_pt(question)
    if not palavras or len(palavras) > 8:
        return False
    return bool(palavras & _VAGUE_INSIGHT_GATILHOS)


def _proximo_tema_vago_insight(phone: str) -> tuple[str, str]:
    """Escolhe o próximo tema da rotação, evitando repetir o último usado
    por esse número."""
    from shared.session_store import get_last_vague_insight_metric, set_last_vague_insight_metric
    ultimo = get_last_vague_insight_metric(phone) if phone else None
    candidatos = [t for t in _VAGUE_INSIGHT_TEMAS if t[0] != ultimo] or _VAGUE_INSIGHT_TEMAS
    escolhido = random.choice(candidatos)
    if phone:
        set_last_vague_insight_metric(phone, escolhido[0])
    return escolhido


async def extract_node(state: State) -> dict:
    """Extrai parâmetros estruturados da pergunta do usuário."""
    question              = state.get("question") or ""
    last_question         = state.get("last_question")
    last_answer           = state.get("last_answer")
    last_extracted_params = state.get("last_extracted_params")
    today                 = date.today().isoformat()

    # Inclui o insight anterior no contexto para follow-ups que referenciam o insight
    phone = state.get("phone") or ""
    pending_insight: str | None = None
    if phone:
        from shared.session_store import get_pending_insight as _gpi
        pi = _gpi(phone)
        if pi:
            pending_insight = pi.get("insight")

    profile_hint = state.get("profile_hint")
    _log("extract", question=question[:80], has_history=bool(last_question), has_insight=bool(pending_insight))

    try:
        params = await _generator.extract_params(
            question=question,
            today=today,
            last_question=last_question,
            last_answer=last_answer,
            last_extracted_params=last_extracted_params,
            pending_insight=pending_insight,
            profile_hint=profile_hint,
        )
        params_dict = params.model_dump()

        # Pedido vago de "insight"/"análise" sem métrica/produto/categoria —
        # rotaciona o tema em vez de sempre cair em faturamento dos últimos 30 dias
        if (
            params_dict.get("metric") == "outro"
            and not params_dict.get("product_filter")
            and not params_dict.get("category_filter")
            and _eh_pedido_vago_insight(question)
        ):
            metric, tema = _proximo_tema_vago_insight(phone)
            params_dict["metric"] = metric
            params_dict["preferred_table"] = "auto"
            params_dict["summary"] = f"Pedido vago de insight — gerar análise sobre {tema}."
            _log("extract", result=f"pedido vago de insight → rotacionado para metric={metric}")

        _log("extract",
             metric=params.metric,
             grain=params.grain,
             period=params.period_type,
             table=params.preferred_table,
             product=params.product_filter,
             category=params.category_filter,
             limit=params.limit,
             summary=params.summary[:60])
        return {"extracted_params": params_dict}
    except Exception as e:
        _log("extract", result=f"ERRO: {e}")
        return {"extracted_params": None}


def _cnpj_digits(cnpj: str) -> str:
    """Retorna apenas os dígitos do CNPJ. Para CNPJs demo (sem dígitos), retorna string vazia."""
    return re.sub(r"\D", "", cnpj or "")


def _is_demo_cnpj(cnpj: str) -> bool:
    """Retorna True para CNPJs sem dígitos (ex: XXXXXXXXXXXXXX)."""
    return bool(cnpj) and not _cnpj_digits(cnpj)


def _build_no_entrada_message(cnpj: str) -> str:
    """Monta mensagem quando não há notas de entrada (certificado vencido/ausente)."""
    cnpj_digits = _cnpj_digits(cnpj)
    cnpj_rev = cnpj_digits[::-1] if cnpj_digits else "00000000000000"

    url_cert = (
        "https://simtech.martins.com.br/"
        f"SIMTECH.CONTRATO/Paginas/Questionarios/Index?1#{cnpj_rev}#1;2;15;16"
    )

    return (
        "Não encontrei notas de entrada para o seu CNPJ.\n\n"
        "Para garantir que as notas de compra sejam capturadas corretamente, "
        "é importante que o certificado digital esteja atualizado.\n\n"
        f"Acesse o portal para atualizar ou validar o seu certificado digital:\n{url_cert}\n\n"
        "Se tiver dúvidas, fale com o nosso suporte:\n"
        "- WhatsApp: (34) 99912-7261\n"
        "- Capitais e regiões metropolitanas: 3003-1266\n"
        "- Outras regiões: 0800-729-5217"
    )


async def cert_check_node(state: State) -> dict:
    """Para metric=gastos, verifica certificado digital antes de gerar SQL."""
    params_dict = state.get("extracted_params") or {}
    metric = params_dict.get("metric", "")
    cnpj = state.get("cnpj") or ""

    # Se não é gastos, passa direto (sem alterar nada)
    if metric != "gastos":
        _log("cert_check", result="SKIP (não é gastos)")
        return {}

    # CNPJ demo (sem dígitos, ex: XXXXXXXXXXXXXX) — pula verificação
    if _is_demo_cnpj(cnpj):
        _log("cert_check", result="SKIP (CNPJ demo sem dígitos)")
        return {}

    _log("cert_check", result="verificando notas de entrada")

    # Passo 1: verifica se existem notas de ENTRADA nos últimos 90 dias
    entrada_sql = (
        f"SELECT COUNT(*) AS QTD "
        f"FROM imaiscatalog.gold_prod.mvp_dados_intermediarios i "
        f"WHERE i.CNPJ = '{cnpj}' "
        f"AND i.TIPO_NOTA = 'ENTRADA' "
        f"AND to_date(i.DATA_EMISSAO) >= date_sub(current_date(), 90)"
    )

    try:
        js = await run_query(entrada_sql)
        rows = parse_rows(js)
        qtd = int((rows or [[0]])[0][0] or 0)

        if qtd > 0:
            _log("cert_check", result=f"{qtd} notas de entrada → prossegue")
            return {}

        # Passo 2: sem notas → verifica o certificado para dar mensagem precisa
        _log("cert_check", result="sem notas de entrada → verificando certificado")
        _digits = _cnpj_digits(cnpj).lstrip("0") or _cnpj_digits(cnpj)
        cert_sql = (
            f"SELECT max(to_date(cert.DATFIMVLDINF)) AS DT_FIM_VALIDADE "
            f"FROM imaiscatalog.bronze_prod.cadcrfclitgv cert "
            f"WHERE cast(cert.NUMCGCCLI as string) LIKE '%{_digits}%'"
        )
        js2 = await run_query(cert_sql)
        cert_rows = parse_rows(js2)
        cert_date_str = str((cert_rows or [[None]])[0][0] or "").strip()

        if not cert_date_str:
            _log("cert_check", result="certificado não consta na base → bloqueia")
            msg = _build_no_entrada_message(cnpj).replace(
                "Não encontrei notas de entrada para o seu CNPJ.",
                "Não encontrei notas de entrada e seu certificado digital não consta na nossa base."
            )
            return {"direct_reply": msg}

        cert_date = date.fromisoformat(cert_date_str)
        today = date.today()
        if cert_date < today:
            _log("cert_check", result=f"certificado VENCIDO em {cert_date} → bloqueia")
            msg = _build_no_entrada_message(cnpj).replace(
                "Não encontrei notas de entrada para o seu CNPJ.",
                f"Não encontrei notas de entrada. Seu certificado digital está vencido desde {cert_date.strftime('%d/%m/%Y')}."
            )
            return {"direct_reply": msg}

        # Cert válido mas sem notas — pode ser delay de processamento
        _log("cert_check", result=f"cert válido até {cert_date}, mas sem notas → bloqueia")
        msg = _build_no_entrada_message(cnpj).replace(
            "Não encontrei notas de entrada para o seu CNPJ.",
            "Não encontrei notas de entrada nos últimos 90 dias, mas seu certificado está válido. Os dados podem estar em processamento."
        )
        return {"direct_reply": msg}

    except Exception as e:
        _log("cert_check", result=f"ERRO: {e} → prossegue mesmo assim")
        return {}


# ── Comandos de estoque por seção (escalável para açougue, padaria, etc.) ──────

# Seções suportadas: chave = nome normalizado (sem acento, minúsculo)
_SECTION_ALIASES: dict[str, str] = {
    "hortifruti": "hortifruti",
    "frutaria":   "hortifruti",
    "frutas":     "hortifruti",
    "verduras":   "hortifruti",
    # futuro: "acougue": "acougue", "padaria": "padaria"
}


def _detect_estoque_command(message: str) -> tuple[str, str] | None:
    """Detecta comandos de estoque por seção.
    Retorna (mode, section) ou None.
    mode: 'relatorio' | 'reposicao'
    section: 'hortifruti' | ... (futuras seções)
    """
    import re
    import unicodedata

    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )

    patterns = [
        (r"relat[oa]rio\s+(?:de\s+|do\s+|da\s+)?(\w+)", "relatorio"),
        (r"repos[ií]cao\s+(?:de\s+|do\s+|da\s+)?(\w+)",  "reposicao"),
        (r"lista\s+(?:de\s+)?repos[ií]cao\s+(?:de\s+|do\s+|da\s+)?(\w+)", "reposicao"),
        (r"o\s+que\s+repor\s+(?:no\s+|na\s+|em\s+)?(\w+)", "reposicao"),
    ]

    for pattern, mode in patterns:
        m = re.search(pattern, msg_norm)
        if m:
            section = _SECTION_ALIASES.get(m.group(1).strip())
            if section:
                return mode, section

    return None


def _is_reposicao_produto_especifico(message: str) -> bool:
    """Identifica pedidos de compra para um item, mesmo sem citar a seção."""
    import unicodedata

    msg = "".join(
        c for c in unicodedata.normalize("NFD", (message or "").lower())
        if unicodedata.category(c) != "Mn"
    )
    patterns = (
        r"\b(?:reposicao|repor|reponha)\s+(?:de\s+|do\s+|da\s+)?([a-z][a-z\s-]{2,})",
        r"\b(?:sugestao\s+(?:de\s+)?compra|quanto\s+comprar)\s+(?:de\s+|do\s+|da\s+)?([a-z][a-z\s-]{2,})",
    )
    generic_terms = {
        "hortifruti", "flv", "lfv", "vlf", "lista", "produtos", "produto",
        "mercado", "estoque", "semana",
    }
    for pattern in patterns:
        match = re.search(pattern, msg)
        if not match:
            continue
        product = match.group(1).strip().split()
        if product and product[0] not in generic_terms:
            return True
    return bool(re.search(
        r"\b(?:reposicao|repor|reponha|sugestao\s+(?:de\s+)?compra|quanto\s+comprar)"
        r"\s+(?:de\s+|do\s+|da\s+)?(?:produto\s+)?\d+\b",
        msg,
    ))


_HORTIFRUTI_REPORT_SQL = """\
WITH
ref AS (
  SELECT max(DATA) AS max_data FROM imaiscatalog.gold_prod.movimentacao_estoque
),
compras_semanais AS (
  SELECT
    m.PRODUTO_CHAVE,
    date_trunc('WEEK', m.DATA) AS SEMANA,
    SUM(
      CASE
        WHEN m.UCOM_ENTRADA <> m.UCOM_SAIDA AND m.FATOR > 1 AND m.FATOR <= 1000
          THEN m.QUANTIDADE_COMPRADA * m.FATOR
        ELSE COALESCE(NULLIF(m.QUANTIDADE_COMPRADA_CONVERTIDA, 0),
                      m.QUANTIDADE_COMPRADA * COALESCE(m.FATOR, 1))
      END
    ) AS QTD_COMPRADA
  FROM imaiscatalog.gold_prod.movimentacao_estoque m
  CROSS JOIN ref
  WHERE m.DATA >= ref.max_data - INTERVAL 84 DAYS
    AND m.QUANTIDADE_COMPRADA > 0
  GROUP BY m.PRODUTO_CHAVE, date_trunc('WEEK', m.DATA)
),
historico_semanal AS (
  SELECT
    PRODUTO_CHAVE,
    AVG(QTD_COMPRADA) AS MEDIA_COMPRA_SEMANAL
  FROM compras_semanais
  GROUP BY PRODUTO_CHAVE
),
ultima_compra AS (
  SELECT
    PRODUTO_CHAVE,
    DATA AS DATA_ULTIMA_COMPRA,
    QUANTIDADE_COMPRADA AS QTD_ULTIMA_COMPRA_ENTRADA,
    CASE
      WHEN UCOM_ENTRADA <> UCOM_SAIDA AND FATOR > 1 AND FATOR <= 1000
        THEN QUANTIDADE_COMPRADA * FATOR
      ELSE COALESCE(NULLIF(QUANTIDADE_COMPRADA_CONVERTIDA, 0),
                    QUANTIDADE_COMPRADA * COALESCE(FATOR, 1))
    END AS QTD_ULTIMA_COMPRA_SAIDA,
    UCOM_ENTRADA,
    FATOR,
    FORNECEDOR,
    ROW_NUMBER() OVER (PARTITION BY PRODUTO_CHAVE ORDER BY DATA DESC) AS rn
  FROM imaiscatalog.gold_prod.movimentacao_estoque
  WHERE QUANTIDADE_COMPRADA > 0
),
embalagens_por_fornecedor AS (
  SELECT DISTINCT
    m.PRODUTO_CHAVE,
    m.FORNECEDOR,
    m.UCOM_ENTRADA,
    m.FATOR
  FROM imaiscatalog.gold_prod.movimentacao_estoque m
  CROSS JOIN ref
  WHERE m.DATA >= ref.max_data - INTERVAL 90 DAYS
    AND m.QUANTIDADE_COMPRADA > 0
    AND m.FORNECEDOR IS NOT NULL
    AND m.UCOM_ENTRADA <> m.UCOM_SAIDA
    AND m.FATOR > 1
    AND m.FATOR <= 1000
),
opcoes_embalagem AS (
  SELECT
    PRODUTO_CHAVE,
    concat_ws(
      '||',
      sort_array(collect_set(concat(UCOM_ENTRADA, '::', cast(FATOR AS string), '::', FORNECEDOR)))
    ) AS OPCOES_EMBALAGEM
  FROM embalagens_por_fornecedor
  GROUP BY PRODUTO_CHAVE
),
consumo_desde_ultima_compra AS (
  SELECT
    m.PRODUTO_CHAVE,
    SUM(COALESCE(m.QUANTIDADE_VENDIDA, 0)) AS CONSUMO_CICLO,
    DATEDIFF(ref.max_data, c.DATA_ULTIMA_COMPRA) + 1 AS DIAS_DESDE_ULTIMA_COMPRA
  FROM imaiscatalog.gold_prod.movimentacao_estoque m
  JOIN ultima_compra c
    ON c.PRODUTO_CHAVE = m.PRODUTO_CHAVE
   AND c.rn = 1
   AND m.DATA >= c.DATA_ULTIMA_COMPRA
  CROSS JOIN ref
  GROUP BY m.PRODUTO_CHAVE, DATEDIFF(ref.max_data, c.DATA_ULTIMA_COMPRA) + 1
),
vendas_semanais AS (
  SELECT
    m.PRODUTO_CHAVE,
    date_trunc('WEEK', m.DATA) AS SEMANA,
    SUM(COALESCE(m.QUANTIDADE_VENDIDA, 0)) AS QTD_VENDIDA
  FROM imaiscatalog.gold_prod.movimentacao_estoque m
  CROSS JOIN ref
  WHERE m.DATA >= ref.max_data - INTERVAL 84 DAYS
  GROUP BY m.PRODUTO_CHAVE, date_trunc('WEEK', m.DATA)
),
media_venda_semanal AS (
  SELECT
    PRODUTO_CHAVE,
    AVG(QTD_VENDIDA) AS MEDIA_VENDA_SEMANAL
  FROM vendas_semanais
  GROUP BY PRODUTO_CHAVE
),
snapshot AS (
  SELECT
    PRODUTO_CHAVE,
    DESC_ENTRADA,
    CPROD_ENTRADA,
    CPROD_SAIDA,
    UCOM_SAIDA,
    ESTOQUE_FINAL AS SALDO_ATUAL,
    ROW_NUMBER() OVER (PARTITION BY PRODUTO_CHAVE ORDER BY DATA DESC) AS rn
  FROM imaiscatalog.gold_prod.movimentacao_estoque
)
SELECT
  s.PRODUTO_CHAVE,
  s.DESC_ENTRADA,
  s.UCOM_SAIDA,
  s.SALDO_ATUAL,
  COALESCE(v.CONSUMO_CICLO, 0) AS CONSUMO_DESDE_ULTIMA_COMPRA,
  COALESCE(mv.MEDIA_VENDA_SEMANAL, 0) AS MEDIA_VENDA_SEMANAL,
  COALESCE(h.MEDIA_COMPRA_SEMANAL, 0) AS MEDIA_COMPRA_SEMANAL,
  c.QTD_ULTIMA_COMPRA_ENTRADA,
  c.QTD_ULTIMA_COMPRA_SAIDA,
  c.UCOM_ENTRADA,
  c.FATOR,
  c.DATA_ULTIMA_COMPRA,
  COALESCE(v.DIAS_DESDE_ULTIMA_COMPRA, 0) AS DIAS_DESDE_ULTIMA_COMPRA,
  s.CPROD_SAIDA AS CODIGO_PRODUTO,
  o.OPCOES_EMBALAGEM,
  c.FORNECEDOR AS FORNECEDOR_ULTIMA_COMPRA
FROM snapshot s
 CROSS JOIN ref
LEFT JOIN consumo_desde_ultima_compra v ON v.PRODUTO_CHAVE = s.PRODUTO_CHAVE
LEFT JOIN media_venda_semanal mv ON mv.PRODUTO_CHAVE = s.PRODUTO_CHAVE
LEFT JOIN historico_semanal h ON h.PRODUTO_CHAVE = s.PRODUTO_CHAVE
LEFT JOIN ultima_compra c ON c.PRODUTO_CHAVE = s.PRODUTO_CHAVE AND c.rn = 1
LEFT JOIN opcoes_embalagem o ON o.PRODUTO_CHAVE = s.PRODUTO_CHAVE
WHERE s.rn = 1
{eligibility_filter}
{product_filter}
ORDER BY (
  GREATEST(
    COALESCE(v.CONSUMO_CICLO, 0),
    COALESCE(mv.MEDIA_VENDA_SEMANAL, 0),
    COALESCE(h.MEDIA_COMPRA_SEMANAL, 0)
  ) - s.SALDO_ATUAL
) DESC"""

_PERDAS_HISTORICO_SQL = """\
SELECT e.CODIGO, e.DESCRICAO, e.UNIDADE_MEDIDA, e.QNT_ESTOQUE, e.DATA_RELATORIO, e.TIPO_OPERACAO
FROM imaiscatalog.silver_prod.estoque_quantum_poc e
JOIN imaiscatalog.gold_prod.dim_cli c ON c.SRK_CLI = e.SRK_CLI
WHERE lpad(cast(c.CNPJ_CPF as string), 14, '0') = '{cnpj}'
  AND {data_filter}
ORDER BY e.DATA_RELATORIO DESC, e.DESCRICAO
LIMIT 200"""


def _parse_perdas_data_filter(question: str) -> tuple[str, str, str, str]:
    """Extrai filtro de data da pergunta.
    Retorna (sql_filter, periodo_label, data_inicio_iso, data_fim_iso).
    """
    import re
    from datetime import date as _d, timedelta as _td

    q = _normalize_pt(question or "")
    today = _d.today()

    if re.search(r'\bhoje\b', q):
        d = today.isoformat()
        return f"DATE(e.DATA_RELATORIO) = '{d}'", "hoje", d, d

    if re.search(r'\bontem\b', q):
        d = (today - _td(days=1)).isoformat()
        return f"DATE(e.DATA_RELATORIO) = '{d}'", "ontem", d, d

    if re.search(r'\b(essa|esta)\s+semana\b', q):
        start = (today - _td(days=today.weekday())).isoformat()
        return f"DATE(e.DATA_RELATORIO) >= '{start}'", "essa semana", start, today.isoformat()

    if re.search(r'\b(esse|este|o)\s+mes\b', q):
        start = today.replace(day=1).isoformat()
        return f"DATE(e.DATA_RELATORIO) >= '{start}'", "este mês", start, today.isoformat()

    # Range: "de DD/MM até DD/MM" ou "entre DD/MM e DD/MM"
    rng = re.search(
        r'(?:de|entre)\s+(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?\s+(?:ate|a|e)\s+(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?',
        q,
    )
    if rng:
        g = rng.groups()
        y1 = int(g[2]) if g[2] else today.year
        y2 = int(g[5]) if g[5] else today.year
        if y1 < 100: y1 += 2000
        if y2 < 100: y2 += 2000
        d1 = _d(y1, int(g[1]), int(g[0]))
        d2 = _d(y2, int(g[4]), int(g[3]))
        return (
            f"DATE(e.DATA_RELATORIO) BETWEEN '{d1.isoformat()}' AND '{d2.isoformat()}'",
            f"{d1.strftime('%d/%m')} até {d2.strftime('%d/%m')}",
            d1.isoformat(), d2.isoformat(),
        )

    # Data específica: DD/MM ou DD/MM/YYYY
    single = re.search(r'\b(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?\b', q)
    if single:
        g = single.groups()
        y = int(g[2]) if g[2] else today.year
        if y < 100: y += 2000
        d = _d(y, int(g[1]), int(g[0]))
        return f"DATE(e.DATA_RELATORIO) = '{d.isoformat()}'", d.strftime('%d/%m/%Y'), d.isoformat(), d.isoformat()

    # Padrão: últimos 7 dias
    start = (today - _td(days=6)).isoformat()
    return f"DATE(e.DATA_RELATORIO) >= '{start}'", "últimos 7 dias", start, today.isoformat()


def _gerar_excel_perdas(rows: list, periodo: str) -> bytes:
    """Gera arquivo Excel (.xlsx) com o histórico de movimentações."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    headers = ["Código", "Produto", "Unidade", "Quantidade", "Data/Hora", "Tipo"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    _TIPO_LABEL = {
        "LIXO": "Descarte", "PERDA": "Perda", "USO_INTERNO": "Uso interno",
        "LANCHE": "Lanche", "BAIXA": "Baixa", "LANCAMENTO": "Lançamento",
    }

    for row_idx, r in enumerate(rows, 2):
        try:
            tipo = str(r[5] or "").upper() if len(r) > 5 and r[5] else ""
            ws.cell(row=row_idx, column=1, value=str(r[0] or ""))
            ws.cell(row=row_idx, column=2, value=str(r[1] or ""))
            ws.cell(row=row_idx, column=3, value=str(r[2] or ""))
            ws.cell(row=row_idx, column=4, value=float(r[3]) if r[3] is not None else 0)
            ws.cell(row=row_idx, column=5, value=str(r[4] or ""))
            ws.cell(row=row_idx, column=6, value=_TIPO_LABEL.get(tipo, tipo.capitalize()))
        except (TypeError, ValueError, IndexError):
            continue

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _gerar_excel_resumo_reposicao(triage: dict) -> bytes:
    """Gera o resumo de compra com uma linha para cada fornecedor e embalagem."""
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sugestao de compra"

    headers = ["PRODUTO", "CODIGO", "FORNECEDOR", "VENDA MEDIA", "SUGESTAO DE COMPRA", "TOTAL"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_index = 2
    for item in _sugestoes_ordenadas_por_giro(triage):
        venda_media = (
            float(item.get("consumo_ciclo") or 0.0) / max(1, int(item.get("dias_ciclo") or 1))
            if item.get("consumo_ciclo")
            else float(item.get("media_sem") or 0.0) / 7
        )
        quantidade = float(item.get("quantidade_sugerida") or item.get("reposicao") or 0.0)
        opcoes: list[tuple[str, float, str]] = []
        for raw in str(item.get("opcoes_embalagem") or "").split("||"):
            try:
                unidade, fator_raw, fornecedor = raw.split("::", 2)
                fator = float(fator_raw)
            except (TypeError, ValueError):
                continue
            if unidade and fornecedor and 1 < fator <= 1000:
                opcoes.append((unidade, fator, fornecedor))

        if not opcoes:
            opcoes = [(item.get("ucom_entrada") or item.get("un") or "UN", 1.0, item.get("fornecedor") or "Não informado")]

        for unidade, fator, fornecedor in sorted(opcoes, key=lambda option: (option[2], option[1])):
            if fator > 1 and unidade != item.get("un"):
                caixas = math.ceil(quantidade / fator)
                sugestao = f"{caixas} {unidade} de {fator:g} {item.get('un', 'UN')}"
                total = f"{caixas * fator:g} {item.get('un', 'UN')}"
            else:
                sugestao = f"{quantidade:.2f} {item.get('un', 'UN')}".replace(".", ",")
                total = sugestao

            ws.cell(row=row_index, column=1, value=item["desc"])
            ws.cell(row=row_index, column=2, value=str(item.get("codigo") or ""))
            ws.cell(row=row_index, column=3, value=fornecedor)
            ws.cell(row=row_index, column=4, value=f"{venda_media:.2f} {item.get('un', 'UN')}/dia".replace(".", ","))
            ws.cell(row=row_index, column=5, value=sugestao)
            ws.cell(row=row_index, column=6, value=total)
            row_index += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(row_index - 1, 1)}"
    for column in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 3, 55)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Busca a cidade do cliente (MUNICIPIO mais frequente nas notas)
_HORTIFRUTI_CITY_SQL = """\
SELECT MUNICIPIO
FROM imaiscatalog.gold_prod.mvp_dados_intermediarios
WHERE CNPJ = '{cnpj}'
  AND MUNICIPIO IS NOT NULL
  AND trim(MUNICIPIO) != ''
GROUP BY MUNICIPIO
ORDER BY COUNT(*) DESC
LIMIT 1"""

# Top produtos hortifruti vendidos na cidade (últimos 90 dias) — para comparação de mix
_HORTIFRUTI_MARKET_SQL = """\
SELECT
  i.DESC_PROD,
  ROUND(SUM(CAST(i.QUANTIDADE_COMPRADA AS DECIMAL(15,2))), 1) AS QTD_TOTAL,
  COUNT(DISTINCT i.CNPJ) AS N_LOJAS
FROM imaiscatalog.gold_prod.mvp_dados_intermediarios i
WHERE i.SECAO LIKE '%HORTIFRUTI%'
  AND i.MUNICIPIO = '{municipio}'
  AND i.TIPO_NOTA = 'SAIDA'
  AND to_date(i.DATA_EMISSAO) >= date_sub(current_date(), 90)
GROUP BY i.DESC_PROD
ORDER BY QTD_TOTAL DESC
LIMIT 40"""


# ── Promoção: candidatos a oferta com preço sugerido vs concorrência ──────────
# Cruza a curva ABCD do cliente (giro) com o preço praticado pela concorrência na
# MESMA UF para os MESMOS produtos (CODIGO_GTIN). Seleciona itens de bom giro
# (curva A/B) onde o preço da loja está acima do mercado — espaço para baixar e
# ainda atrair tráfego. Preço sugerido = ~5% abaixo da média do mercado.
_PROMOCAO_SQL = """\
WITH loja AS (
  SELECT
    i.CODIGO_GTIN,
    MAX(i.DESC_PROD)                                   AS DESC_PROD,
    MAX(i.UNIDADE)                                     AS UNIDADE,
    MAX(i.UF)                                          AS UF,
    SUM(CAST(i.QUANTIDADE_COMPRADA AS DECIMAL(15,2)))  AS QTD_LOJA,
    AVG(CAST(i.VALOR_UNITARIO AS DOUBLE))              AS PRECO_LOJA
  FROM imaiscatalog.gold_prod.mvp_dados_intermediarios i
  WHERE i.CNPJ = '{cnpj}'
    AND i.TIPO_NOTA = 'SAIDA'
    AND i.CODIGO_GTIN IS NOT NULL AND trim(i.CODIGO_GTIN) <> ''
    AND CAST(i.VALOR_UNITARIO AS DOUBLE) > 0
    AND to_date(i.DATA_EMISSAO) BETWEEN date_sub(current_date(), 30) AND date_sub(current_date(), 1)
  GROUP BY i.CODIGO_GTIN
),
curva AS (
  SELECT CODIGO_GTIN,
         MAX(CURVA_CLIENTE) AS CURVA_CLIENTE,
         MAX(CURVA_MERCADO) AS CURVA_MERCADO
  FROM imaiscatalog.gold_prod.nova_mvp_curva_abcd
  WHERE CNPJ = '{cnpj}'
  GROUP BY CODIGO_GTIN
),
mercado AS (
  SELECT
    m.CODIGO_GTIN,
    AVG(CAST(m.VALOR_UNITARIO AS DOUBLE))  AS PRECO_MERCADO_MEDIO,
    MIN(CAST(m.VALOR_UNITARIO AS DOUBLE))  AS PRECO_MERCADO_MIN,
    COUNT(DISTINCT m.CNPJ)                 AS N_LOJAS_MERCADO
  FROM imaiscatalog.gold_prod.mvp_dados_intermediarios m
  JOIN loja l ON m.CODIGO_GTIN = l.CODIGO_GTIN AND m.UF = l.UF
  WHERE m.CNPJ <> '{cnpj}'
    AND m.TIPO_NOTA = 'SAIDA'
    AND CAST(m.VALOR_UNITARIO AS DOUBLE) > 0
    AND to_date(m.DATA_EMISSAO) BETWEEN date_sub(current_date(), 30) AND date_sub(current_date(), 1)
  GROUP BY m.CODIGO_GTIN
)
SELECT
  l.DESC_PROD,
  l.UNIDADE,
  COALESCE(c.CURVA_CLIENTE, 'N/D')                     AS CURVA_CLIENTE,
  COALESCE(c.CURVA_MERCADO, 'N/D')                     AS CURVA_MERCADO,
  ROUND(l.QTD_LOJA, 1)                                 AS QTD_VENDIDA,
  ROUND(l.PRECO_LOJA, 2)                               AS PRECO_ATUAL,
  ROUND(mk.PRECO_MERCADO_MEDIO, 2)                     AS PRECO_MERCADO_MEDIO,
  ROUND(mk.PRECO_MERCADO_MIN, 2)                       AS PRECO_MERCADO_MIN,
  ROUND(mk.PRECO_MERCADO_MEDIO * 0.95, 2)              AS PRECO_SUGERIDO,
  mk.N_LOJAS_MERCADO                                   AS N_LOJAS,
  ROUND(100.0 * (l.PRECO_LOJA - mk.PRECO_MERCADO_MEDIO) / NULLIF(mk.PRECO_MERCADO_MEDIO, 0), 1) AS GAP_PCT
FROM loja l
JOIN mercado mk  ON l.CODIGO_GTIN = mk.CODIGO_GTIN
LEFT JOIN curva c ON l.CODIGO_GTIN = c.CODIGO_GTIN
WHERE mk.N_LOJAS_MERCADO >= 2
  AND l.PRECO_LOJA > mk.PRECO_MERCADO_MEDIO * 0.95
ORDER BY
  CASE WHEN c.CURVA_CLIENTE IN ('CURVA A', 'CURVA B') THEN 0 ELSE 1 END,
  (l.PRECO_LOJA - mk.PRECO_MERCADO_MEDIO) DESC,
  l.QTD_LOJA DESC
LIMIT 12"""


_PROMOCAO_KEYWORDS = [
    "colocar em promocao", "produtos em promocao", "itens em promocao",
    "em promocao", "promoção", "promocao", "em oferta", "colocar em oferta",
    "produtos para promover", "o que promover", "que promover", "sugestao de promocao",
    "preco promocional", "promoção essa semana", "fazer promocao", "fazer uma promocao",
]


def _is_promocao_question(message: str) -> bool:
    """Detecta pedidos de sugestão de produtos para promoção / preço promocional."""
    import unicodedata
    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )
    return any(kw in msg_norm for kw in _PROMOCAO_KEYWORDS)


def _is_perdas_question(message: str) -> bool:
    """Detecta pedidos de histórico de movimentações/perdas do estoque manual."""
    import unicodedata
    msg = (message or "").strip().lower()
    msg_norm = "".join(
        c for c in unicodedata.normalize("NFD", msg)
        if unicodedata.category(c) != "Mn"
    )
    _PERDAS_KEYWORDS = (
        "historico de perdas", "relatorio de perdas", "historico de movimentacoes",
        "movimentacoes manuais", "minhas perdas", "perdas do estoque",
        "historico do estoque", "movimentacoes do estoque", "historico de baixas",
    )
    return any(kw in msg_norm for kw in _PERDAS_KEYWORDS)


def _fmt_perdas_block(rows: list, periodo: str = "últimos 7 dias") -> str:
    """Formata o histórico de movimentações manuais para WhatsApp."""
    from datetime import datetime as _dt

    def _fmt_data(raw) -> str:
        if not raw:
            return "—"
        try:
            return _dt.fromisoformat(str(raw).replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(raw)[:16]

    _TIPO_LABEL = {
        "LIXO":        "Descarte",
        "PERDA":       "Perda",
        "USO_INTERNO": "Uso interno",
        "LANCHE":      "Lanche",
        "BAIXA":       "Baixa",
        "LANCAMENTO":  "Lançamento",
    }

    if not rows:
        return f"Nenhuma movimentação registrada no estoque manual ({periodo})."

    lines = [f"📋 *Histórico de Movimentações — {periodo}*", ""]
    for r in rows:
        try:
            produto = str(r[1] or "—")
            un      = str(r[2] or "") if len(r) > 2 and r[2] else ""
            qtd     = r[3] if len(r) > 3 else None
            data    = _fmt_data(r[4] if len(r) > 4 else None)
            tipo    = str(r[5] or "").upper() if len(r) > 5 and r[5] else ""
            qtd_fmt  = f"{float(qtd):g} {un}".strip() if qtd is not None else "—"
            tipo_fmt = _TIPO_LABEL.get(tipo, tipo.capitalize()) if tipo else "—"
            lines.append(f"*{produto}*")
            lines.append(f"{qtd_fmt}  •  {tipo_fmt}  •  {data}")
            lines.append("")
        except (TypeError, ValueError, IndexError):
            continue

    return "\n".join(lines).strip()


def _fmt_promocao_block(rows: list, cols: list) -> str:
    """Formata os candidatos a promoção em um bloco de texto para o LLM."""
    cl = [str(c).lower() for c in cols]

    def _idx(name: str) -> int:
        try:
            return cl.index(name)
        except ValueError:
            return -1

    i_desc  = _idx("desc_prod")
    i_un    = _idx("unidade")
    i_cc    = _idx("curva_cliente")
    i_cm    = _idx("curva_mercado")
    i_qtd   = _idx("qtd_vendida")
    i_patu  = _idx("preco_atual")
    i_pmed  = _idx("preco_mercado_medio")
    i_pmin  = _idx("preco_mercado_min")
    i_psug  = _idx("preco_sugerido")
    i_nloj  = _idx("n_lojas")
    i_gap   = _idx("gap_pct")

    def _g(r, idx):
        return r[idx] if 0 <= idx < len(r) else None

    def _money(v):
        try:
            return f"R$ {float(v):.2f}".replace(".", ",")
        except (TypeError, ValueError):
            return "—"

    linhas = []
    for r in rows:
        desc = str(_g(r, i_desc) or "").strip()
        if not desc:
            continue
        un   = str(_g(r, i_un) or "").strip()
        linhas.append(
            f"- {desc} ({un}) | curva loja: {_g(r, i_cc)} / curva mercado: {_g(r, i_cm)} | "
            f"qtd vendida: {_g(r, i_qtd)} | preço atual: {_money(_g(r, i_patu))} | "
            f"média mercado: {_money(_g(r, i_pmed))} | menor mercado: {_money(_g(r, i_pmin))} | "
            f"PREÇO_SUGERIDO: {_money(_g(r, i_psug))} | concorrentes comparados: {_g(r, i_nloj)} | "
            f"gap vs mercado: {_g(r, i_gap)}%"
        )
    return "\n".join(linhas)



def _triage_hortifruti(rows: list, include_all_products: bool = False) -> dict:
    """
      Classifica reposição pelo consumo desde a última compra e pelo histórico semanal.

    Colunas do SQL:
      r[0]=PRODUTO_CHAVE, r[1]=DESC_ENTRADA, r[2]=UCOM_SAIDA,
        r[3]=SALDO_ATUAL (ESTOQUE_FINAL),
        r[4]=CONSUMO_DESDE_ULTIMA_COMPRA,
        r[5]=MEDIA_VENDA_SEMANAL, r[6]=MEDIA_COMPRA_SEMANAL,
        r[7]=QTD_ULTIMA_COMPRA_ENTRADA, r[8]=QTD_ULTIMA_COMPRA_SAIDA,
        r[9]=UCOM_ENTRADA, r[10]=FATOR, r[11]=DATA_ULTIMA_COMPRA,
        r[12]=DIAS_DESDE_ULTIMA_COMPRA, r[13]=CODIGO_PRODUTO,
        r[14]=OPCOES_EMBALAGEM, r[15]=FORNECEDOR_ULTIMA_COMPRA
    """
    critico, gerenciados = [], []
    all_classificados: list[tuple[str, dict]] = []

    for r in rows:
        try:
            # DESC_ENTRADA (r[1]) costuma vir NULL na tabela gold; PRODUTO_CHAVE (r[0])
            # é o nome limpo de exibição (ex: "BANANA PRATA KG").
            descricao    = str(r[1] or r[0] or "").strip()
            ucom_saida   = str(r[2] or "UN").strip() or "UN"
            saldo_atual  = float(r[3]) if r[3] is not None else 0.0
            consumo_ciclo = float(r[4]) if len(r) > 4 and r[4] is not None else 0.0
            media_venda   = float(r[5]) if len(r) > 5 and r[5] is not None else 0.0
            media_compra  = float(r[6]) if len(r) > 6 and r[6] is not None else 0.0
            qtd_entrada   = float(r[7]) if len(r) > 7 and r[7] is not None else None
            qtd_saida     = float(r[8]) if len(r) > 8 and r[8] is not None else None
            ucom_entrada  = str(r[9] or ucom_saida).strip() if len(r) > 9 and r[9] else ucom_saida
            fator         = float(r[10]) if len(r) > 10 and r[10] else None
            data_compra   = r[11] if len(r) > 11 else None
            dias_ciclo    = max(1, int(float(r[12]))) if len(r) > 12 and r[12] is not None else 1
            codigo        = str(r[13] or "").strip() if len(r) > 13 else ""
            opcoes_embalagem = str(r[14] or "").strip() if len(r) > 14 else ""
            fornecedor    = str(r[15] or "").strip() if len(r) > 15 else ""
        except (TypeError, ValueError, IndexError):
            continue

        if not descricao or " XX" in descricao.upper():
            continue

        produto = descricao.upper()
        if "BANANA" in produto:
            cobertura_dias = 7
        elif any(nome in produto for nome in ("ALFACE", "RUCULA", "COUVE", "CHEIRO VERDE", "ESPINAFRE")):
            cobertura_dias = 3
        elif any(nome in produto for nome in ("MORANGO", "FRAMBOESA", "AMORA", "COGUMELO")):
            cobertura_dias = 4
        elif any(nome in produto for nome in ("BATATA", "ABOBORA", "CEBOLA", "ALHO", "CENOURA")):
            cobertura_dias = 7
        else:
            cobertura_dias = 5

        # O saldo negativo não é abatido da compra. A referência é o giro diário
        # com cobertura compatível com a perecibilidade, não a venda acumulada.
        # O ciclo desde a última compra reflete a demanda atual. A média semanal
        # só é usada quando ainda não existe movimentação suficiente no ciclo;
        # compra histórica não é giro e não pode inflar a sugestão.
        giro_diario = (
            consumo_ciclo / dias_ciclo
            if consumo_ciclo > 0
            else media_venda / 7
        )
        estoque_estimado = (qtd_saida - consumo_ciclo) if qtd_saida is not None else None
        estoque_suficiente = estoque_estimado is not None and estoque_estimado >= giro_diario * 3
        if estoque_suficiente and not include_all_products:
            continue

        consumo_referencia = giro_diario * cobertura_dias
        com_margem = consumo_referencia * 1.10
        compra_em_embalagens = (
            fator
            and fator > 1
            and fator <= 1000
            and ucom_entrada != ucom_saida
        )
        if estoque_suficiente:
            embalagens = None
            reposicao = 0.0
            com_margem = 0.0
        elif compra_em_embalagens:
            embalagens = math.ceil(com_margem / fator)
            reposicao = round(embalagens * fator, 3)
        elif ucom_saida.upper() == "UN":
            com_margem = float(math.floor(com_margem))
            embalagens = int(com_margem)
            reposicao = float(embalagens)
        else:
            embalagens = None
            reposicao = round(com_margem, 3)

        item = {
            "desc":         descricao,
            "codigo":       codigo,
            "un":           ucom_saida,
            "ucom_entrada": ucom_entrada,
            "saldo_atual":  saldo_atual,
            "consumo_ciclo": consumo_ciclo,
            "estoque_estimado": estoque_estimado,
            "media_compra":  media_compra,
            "qtd_entrada":   qtd_entrada,
            "qtd_comprada":  qtd_saida,
            "data_compra":   data_compra,
            "dias_ciclo":    dias_ciclo,
            "cobertura_dias": cobertura_dias,
            "reposicao":    reposicao,
            "quantidade_sugerida": com_margem,
            "embalagens":   embalagens,
            "opcoes_embalagem": opcoes_embalagem,
            "fornecedor":  fornecedor,
            "movimentacao": consumo_referencia,
            "fator":        fator,
            "qnt_atu":      saldo_atual,
            "media_sem":    media_venda,
            "semanas":      999,
            "variacao":     0.0,
            "sugestao":     reposicao,
            "series":       [],
            "motivo":       (
                "saldo suficiente para o giro atual; não precisa comprar agora"
                if estoque_suficiente else (
                    "giro alto, cobertura insuficiente para perecível"
                    if saldo_atual <= 0 else "giro ativo, reposição preventiva para perecível"
                )
            ),
        }

        if saldo_atual < 0:
            item["status"] = "critico"
            critico.append(item)
            all_classificados.append(("critico", item))
        else:
            item["status"] = "ok"
            gerenciados.append(item)
            all_classificados.append(("gerenciados", item))

    # maior déficit primeiro
    critico.sort(key=lambda x: x["reposicao"], reverse=True)
    gerenciados.sort(key=lambda x: x["reposicao"], reverse=True)

    MAX_CRITICO     = 15
    MAX_GERENCIADOS = 15

    return {
        "critico":                critico[:MAX_CRITICO],
        "comprar_mais":           [],
        "comprar_menos":          [],
        "gerenciados":            gerenciados[:MAX_GERENCIADOS],
        "parados":                [],
        "total":                  len(critico) + len(gerenciados),
        "n_critico":              len(critico),
        "n_comprar_mais":         0,
        "n_comprar_menos":        0,
        "n_gerenciados":          len(gerenciados),
        "n_parados":              0,
        "omitidos_critico":       max(0, len(critico)     - MAX_CRITICO),
        "omitidos_comprar_mais":  0,
        "omitidos_comprar_menos": 0,
        "omitidos_gerenciados":   max(0, len(gerenciados) - MAX_GERENCIADOS),
        "omitidos_parados":       0,
        "_all_urgentes":          critico,
        "_all_comprar_menos":     [],
        "_all_parados":           [],
        "_all_classificados":     all_classificados,
    }


def _fmt_triage_block(triage: dict) -> str:
    """Serializa giro, compra e embalagem para ajuste de cobertura pela IA."""
    lines = [
        "PRODUTO | UNIDADE_SAIDA | UNIDADE_COMPRA | FATOR_EMBALAGEM | SALDO_FINAL | "
        "MOVIMENTACAO_DESDE_ULTIMA_COMPRA | DIAS_DESDE_ULTIMA_COMPRA | GIRO_DIA | "
        "MEDIA_VENDA_SEMANAL | MEDIA_COMPRA_SEMANAL | COBERTURA_BASE_DIAS | "
        "DATA_ULTIMA_COMPRA | ULTIMA_COMPRA | SUGESTAO_BASE",
        "",
    ]

    def _row(item: dict) -> str:
        dias = item.get("dias_ciclo") or 1
        giro_dia = (item.get("consumo_ciclo") or 0.0) / dias
        ultima_compra = (
            f"{item['qtd_entrada']:g} {item['ucom_entrada']} ({item['qtd_comprada']:g} {item['un']})"
            if item.get("qtd_entrada") and item.get("qtd_comprada")
            else "sem registro"
        )
        return (
            f"{item['desc']} | {item['un']} | {item['ucom_entrada']} | {item.get('fator') or 1:g} | "
            f"{item['saldo_atual']:g} | {item['consumo_ciclo']:g} | {dias} | {giro_dia:g} | "
            f"{item['media_sem']:g} | {item['media_compra']:g} | {item['cobertura_dias']} | "
            f"{item.get('data_compra') or 'sem data'} | "
            f"{ultima_compra} | "
            f"{item['reposicao']:g} {item['un']}"
        )

    lines.append("[REPOSIÇÃO URGENTE]")
    lines += [_row(i) for i in triage["critico"]] or ["(nenhum)"]
    if triage.get("omitidos_critico"):
        lines.append(f"  ...e mais {triage['omitidos_critico']} produto(s) urgentes.")

    lines.append("\n[SALDO OK]")
    lines += [_row(i) for i in triage["gerenciados"]] or ["(nenhum)"]
    if triage.get("omitidos_gerenciados"):
        lines.append(f"  ...e mais {triage['omitidos_gerenciados']} produto(s) com saldo OK.")

    return "\n".join(lines)


def _fmt_reposicao_block(
    triage: dict,
    show_all: bool = False,
    only_descs: set[str] | None = None,
    only_codes: set[str] | None = None,
    display_items: list[dict] | None = None,
    title: str = "🛒 sugestao de compra — Hortifruti",
    fornecedor_filter: str | None = None,
) -> str:
    """Formata a compra na unidade de entrada, sem deixar saldo negativo inflar a sugestão.
    only_descs: quando preenchido, exibe SOMENTE os produtos com essas descrições
    (busca na lista completa, sem o corte de top-15)."""
    if only_codes:
        all_items   = triage.get("_all_classificados") or []
        criticos    = [i for b, i in all_items if b == "critico" and str(i.get("codigo") or "") in only_codes]
        gerenciados = [i for b, i in all_items if b == "gerenciados" and str(i.get("codigo") or "") in only_codes]
    elif only_descs:
        all_items   = triage.get("_all_classificados") or []
        criticos    = [i for b, i in all_items if b == "critico"     and i["desc"] in only_descs]
        gerenciados = [i for b, i in all_items if b == "gerenciados" and i["desc"] in only_descs]
    else:
        criticos    = list(triage.get("_all_urgentes") or triage["critico"])
        gerenciados = list(triage.get("gerenciados") or [])

    def _n(value: float) -> str:
        return f"{value:.2f}".replace(".", ",")

    def _format_date(value) -> str:
        if not value:
            return "—"
        raw = str(value).split("T", 1)[0].split(" ", 1)[0]
        try:
            year, month, day = raw.split("-")
            return f"{day}/{month}/{year}"
        except ValueError:
            return raw

    def _embalagens(qtd: float, fator, un_saida: str, un_entrada: str) -> str:
        if not fator or fator <= 1 or fator > 1000 or un_saida == un_entrada:
            return ""
        return f" (~{math.ceil(qtd / fator)} {un_entrada} de {_n(fator)} {un_saida})"

    def _opcoes_de_compra(item: dict, un_saida: str, un_entrada: str) -> str:
        quantidade = float(item.get("quantidade_sugerida") or item.get("reposicao") or 0.0)
        if quantidade <= 0:
            return "Não precisa comprar agora"
        opcoes: list[tuple[str, float]] = []
        for raw in str(item.get("opcoes_embalagem") or "").split("||"):
            try:
                unidade, fator_raw, fornecedor = raw.split("::", 2)
                fator = float(fator_raw)
            except (TypeError, ValueError):
                continue
            if fornecedor_filter and _normalize_pt(fornecedor) != _normalize_pt(fornecedor_filter):
                continue
            if unidade and 1 < fator <= 1000:
                opcoes.append((unidade, fator))

        if not opcoes:
            return f"{_n(quantidade)} {un_saida}{_embalagens(quantidade, item.get('fator'), un_saida, un_entrada)}"

        opcoes.sort(key=lambda opcao: opcao[1])
        embalagens = [
            f"{math.ceil(quantidade / fator)} {unidade} de {_n(fator)} {un_saida}"
            for unidade, fator in opcoes
        ]
        return f"~{_n(quantidade)} {un_saida} ({' | '.join(embalagens)})"

    def _ultima_compra(item: dict, un_saida: str, un_entrada: str) -> str:
        qtd_entrada = item.get("qtd_entrada")
        qtd_saida = item.get("qtd_comprada")
        data = _format_date(item.get("data_compra"))
        if not qtd_saida:
            return f"Última compra ({data}): —"
        if qtd_entrada and un_entrada != un_saida:
            compra = f"{_n(qtd_entrada)} {un_entrada} ({_n(qtd_saida)} {un_saida})"
        else:
            compra = f"{_n(qtd_saida)} {un_saida}"
        return f"Última compra ({data}): {compra}"

    def _item_lines(item: dict, prefix: str) -> list[str]:
        un_saida = item.get("un", "KG")
        un_entrada = item.get("ucom_entrada", un_saida)
        reposicao = item.get("reposicao") or 0.0
        consumo = item.get("consumo_ciclo") or 0.0
        dias = item.get("dias_ciclo") or 1
        cobertura = item.get("cobertura_dias") or 0
        giro_dia = consumo / dias
        reposicao_fmt = _opcoes_de_compra(item, un_saida, un_entrada)
        codigo = item.get("codigo") or "não informado"
        return [
            f"{prefix} {item['desc']} (codigo: {codigo})",
            f"Sugestão de compra: {reposicao_fmt}",
            f"Venda desde última compra: {_n(consumo)} {un_saida}",
            f"Venda média desde última compra: {_n(giro_dia)} {un_saida}/dia ({dias} dia(s))",
            f"Tempo médio de vida do produto: {cobertura} dia(s)",
            _ultima_compra(item, un_saida, un_entrada),
            "",
        ]

    lines = [title, ""]
    if display_items is not None:
        for item in display_items:
            lines.extend(_item_lines(item, "🔴" if item.get("status") == "critico" else "•"))
        return "\n".join(lines).strip()

    for item in criticos:
        lines.extend(_item_lines(item, "🔴"))

    if not criticos and not only_descs:
        lines.extend(["Nenhum produto com sugestão de compra urgente no momento. 🎉", ""])

    for item in gerenciados:
        lines.extend(_item_lines(item, "•"))

    return "\n".join(lines).strip()


def _sugestoes_ordenadas_por_giro(triage: dict) -> list[dict]:
    """Ordena todas as sugestões pelo ritmo diário de venda, maior primeiro."""
    def giro_diario(item: dict) -> float:
        consumo = float(item.get("consumo_ciclo") or 0.0)
        if consumo > 0:
            return consumo / max(1, int(item.get("dias_ciclo") or 1))
        return float(item.get("media_sem") or 0.0) / 7

    return sorted(
        (item for _, item in triage.get("_all_classificados") or []),
        key=giro_diario,
        reverse=True,
    )


def _fornecedores_do_item(item: dict) -> set[str]:
    fornecedores: set[str] = set()
    for raw in str(item.get("opcoes_embalagem") or "").split("||"):
        try:
            _unidade, _fator, fornecedor = raw.split("::", 2)
        except ValueError:
            continue
        if fornecedor:
            fornecedores.add(fornecedor)
    if not fornecedores and item.get("fornecedor"):
        fornecedores.add(str(item["fornecedor"]))
    return fornecedores


def _fornecedores_solicitados(question: str, triage: dict) -> list[str]:
    """Encontra fornecedores citados na solicitação, aceitando pequenas variações."""
    fornecedores = sorted({
        fornecedor
        for _, item in triage.get("_all_classificados") or []
        for fornecedor in _fornecedores_do_item(item)
    })
    question_norm = _normalize_pt(question)
    if "por fornecedor" in question_norm:
        return fornecedores

    palavras_ignorar = {
        "reposicao", "hortifruti", "sugestao", "compra", "comprar", "lista",
        "produto", "produtos", "fornecedor", "fornecedores", "para", "com",
    }
    termos = {
        termo for termo in re.findall(r"[a-z0-9]+", question_norm)
        if len(termo) >= 4 and termo not in palavras_ignorar
    }
    matches: list[str] = []
    for fornecedor in fornecedores:
        palavras_fornecedor = re.findall(r"[a-z0-9]+", _normalize_pt(fornecedor))
        if any(
            termo == palavra or SequenceMatcher(None, termo, palavra).ratio() >= 0.85
            for termo in termos
            for palavra in palavras_fornecedor
            if len(palavra) >= 4
        ):
            matches.append(fornecedor)
    return matches


def _produtos_exibidos_na_sugestao(answer: str) -> set[str]:
    """Extrai os nomes dos produtos já enviados em uma página de sugestão."""
    return set(re.findall(r"^(?:🔴|•)\s+(.+?)\s+\(codigo:", answer or "", re.MULTILINE))


def _fmt_extra_categorias_block(triage: dict, max_each: int = 8) -> str:
    """Bloco extra (montado direto, sem LLM, para evitar qualquer invenção de número)
    com produtos de saída lenta e parados — usado em "me dê mais/outros itens",
    já que a lista de reposição sozinha não cresce além do total de produtos urgentes."""
    lines: list[str] = []

    lentos = list(triage.get("_all_comprar_menos") or [])[:max_each]
    if lentos:
        lines.append("📉 *Saída lenta — considere revisar antes de comprar mais:*")
        lines.append("")
        for item in lentos:
            sem_s = f"~{item['semanas']:.0f} semana(s) de estoque" if item["semanas"] < 900 else "estoque parado, sem giro"
            lines.append(
                f"*{item['desc']}* — Estoque: {item['qnt_atu']} {item['un']}  |  "
                f"Média: {item['media_sem']:.1f} {item['un']}/sem  |  {sem_s}"
            )
        lines.append("")

    parados = list(triage.get("_all_parados") or [])[:max_each]
    if parados:
        lines.append("🛑 *Parados — sem saída nos últimos 30 dias:*")
        lines.append("")
        for item in parados:
            lines.append(f"*{item['desc']}* — Estoque: {item['qnt_atu']} {item['un']}")
        lines.append("")

    return "\n".join(lines).strip()


# ── Clima e FLV ───────────────────────────────────────────────────────────────
_CLIMA_KEYWORDS = (
    "tempo", "clima", "temperatura", "calor", "frio", "quente", "faz calor",
    "faz frio", "graus", "chuva", "chuvoso", "chover", "ensolarado", "nublado",
    "seco", "umidade", "vento",
)


def _eh_pergunta_clima(question: str) -> bool:
    """Detecta perguntas que cruzam clima/temperatura com decisão de compra de FLV."""
    q = _normalize_pt(question)
    return any(kw in q for kw in _CLIMA_KEYWORDS)


async def _buscar_clima_cidade(cidade: str) -> dict | None:
    """Busca clima atual da cidade via wttr.in (API pública, sem chave).
    Retorna dict com temp_c, sensacao_c, umidade, condicao, max_c, min_c — ou None."""
    import httpx as _httpx
    from urllib.parse import quote as _quote
    try:
        cidade_norm = _normalize_pt(cidade.strip())
        url = f"https://wttr.in/{_quote(cidade_norm)}?format=j1"
        async with _httpx.AsyncClient(timeout=8.0) as cli:
            resp = await cli.get(url, headers={"Accept": "application/json", "User-Agent": "curl/7.0"})
        if resp.status_code != 200:
            return None
        data = resp.json()
        cc = (data.get("current_condition") or [{}])[0]
        w  = (data.get("weather") or [{}])[0]
        return {
            "temp_c":     int(cc.get("temp_C", 0)),
            "sensacao_c": int(cc.get("FeelsLikeC", 0)),
            "umidade":    int(cc.get("humidity", 0)),
            "condicao":   ((cc.get("weatherDesc") or [{}])[0]).get("value", ""),
            "max_c":      int(w.get("maxtempC", 0)),
            "min_c":      int(w.get("mintempC", 0)),
        }
    except Exception as _e:
        _log("clima", result=f"AVISO wttr.in: {_e}")
        return None


def _fmt_clima_estoque_block(triage: dict) -> str:
    """Bloco compacto de estoque disponível para a análise de clima —
    top 35 produtos com estoque > 0, ordenados por giro semanal (maior primeiro)."""
    todos = [item for _, item in (triage.get("_all_classificados") or []) if item["qnt_atu"] > 0]
    todos.sort(key=lambda x: -(x.get("media_sem") or 0))
    linhas = ["PRODUTO | Estoque | Unidade | Média/sem | Situação"]
    for p in todos[:35]:
        bucket = p.get("bucket", "")
        linhas.append(
            f"{p['desc']} | {p['qnt_atu']} {p['un']} | {p.get('media_sem', 0):.1f} {p['un']}/sem"
            f" | {_BUCKET_LABELS.get(bucket, bucket)}"
        )
    return "\n".join(linhas)


_PARADOS_KEYWORDS = (
    "parado", "parados", "parou", "pararam", "encalhado", "encalhados",
    "sem sair", "sem vender", "sem venda", "sem giro", "sem saida",
    "nao vende", "nao saem", "nao estao saindo", "nao esta saindo",
)


def _eh_pergunta_parados(question: str) -> bool:
    """Detecta perguntas sobre quais produtos do hortifruti estão parados/sem giro —
    devem ser respondidas com a classificação 'parado' já calculada no triage
    (série semanal de 3 meses), e NUNCA pelo fluxo genérico de data_query, que
    não tem essa noção e pode gerar números contraditórios com o resto da resposta."""
    q = _normalize_pt(question)
    return any(kw in q for kw in _PARADOS_KEYWORDS)


_CONTINUACAO_PALAVRAS = {
    "mais", "outro", "outros", "outra", "outras",
    "continua", "continue", "resto", "restante", "demais", "falta", "faltam",
}


def _eh_continuacao_generica(question: str) -> bool:
    """Detecta follow-ups curtos e genéricos tipo "me mostre mais", "e os outros",
    "tem mais?" — usados pelo cliente para PAGINAR a resposta anterior, seja ela
    qual for. Diferente de "reposicao_mais_hortifruti" (que assume que a lista
    anterior era de reposição), esse helper não presume o assunto — quem decide
    o que continuar é o chamador, olhando o que foi perguntado/respondido antes."""
    palavras = _words_pt(question)
    return bool(palavras) and bool(palavras & _CONTINUACAO_PALAVRAS) and len(palavras) <= 6


def _normalize_pt(s: str) -> str:
    """Remove acentos e baixa a caixa para comparação tolerante de texto em português."""
    import unicodedata as _ud
    s = (s or "").lower().strip()
    return "".join(c for c in _ud.normalize("NFD", s) if _ud.category(c) != "Mn")


def _words_pt(s: str) -> set[str]:
    """Extrai palavras (alfanuméricas, sem acento/pontuação) de um texto em português.
    Necessário para que "alface?" ou "rabanete," combinem com "ALFACE"/"RABANETE"."""
    import re as _re
    return set(_re.findall(r"[a-z0-9]+", _normalize_pt(s)))


# Palavras genéricas que não ajudam a identificar um produto específico —
# evita que perguntas como "devo comprar mais hoje?" disparem uma busca por produto.
_PRODUTO_STOPWORDS = {
    "comprar", "compra", "compras", "vender", "venda", "vendas", "estoque",
    "produto", "produtos", "quero", "queria", "preciso", "devo", "tenho",
    "mais", "menos", "hoje", "agora", "essa", "esse", "esta", "este",
    "qual", "quais", "como", "para", "saber", "fazer", "pode", "podia",
    "ainda", "sobre", "aquele", "aquela", "vamos", "minha", "minhas",
    "meu", "meus", "loja", "mercado", "fala", "falar", "diga", "explica",
    "explique", "relatorio", "reposicao", "lista", "hortifruti", "tambem",
    "sera", "vale", "pena", "semana", "semanas", "nesse", "nessa", "desse",
    "dessa", "sera", "muito", "pouco",
}

# Rótulos exibidos para cada classificação ao responder sobre um produto específico
_BUCKET_LABELS = {
    "critico":     "🔴 CRÍTICO — compra da semana já esgotada, repor com urgência",
    "gerenciados": "✅ OK — ainda há saldo da última compra, sem urgência",
}


def _product_codes_in_question(question: str) -> set[str]:
    """Extrai códigos numéricos informados pelo cliente."""
    return set(re.findall(r"\d+", question or ""))


def _termos_produto_para_like(question: str) -> list[str]:
    """Extrai termos seguros para compor filtros LIKE por produto."""
    return sorted({
        word.upper()
        for word in _words_pt(question)
        if len(word) >= 4 and word not in _PRODUTO_STOPWORDS and not word.isdigit()
    })


def _match_produtos_triage(question: str, triage: dict, max_matches: int = 100) -> list[tuple[str, dict]]:
    """Localiza, na triagem já calculada, produtos citados na pergunta do usuário.

    Busca tolerante a acentos/maiúsculas por palavras significativas (>=4 letras,
    fora da stoplist) que aparecem tanto na pergunta quanto na descrição do produto.
    Usada para responder diretamente sobre um item específico em vez de devolver
    a lista/relatório completo (ex: "devo comprar mais almeirão?").
    """
    q_codes = _product_codes_in_question(question)
    q_words = {w for w in _words_pt(question) if len(w) >= 4} - _PRODUTO_STOPWORDS
    if not q_words and not q_codes:
        return []

    # IMPORTANTE: usar a lista completa (sem corte de exibição) — caso contrário um
    # produto fora do top-N de cada bucket (ex: loja com 800+ itens) nunca seria encontrado.
    matches: list[tuple[str, dict]] = []
    seen: set[str] = set()
    for bucket_name, item in triage.get("_all_classificados") or []:
        if item["desc"] in seen:
            continue
        item_code = str(item.get("codigo") or "").strip()
        desc_words = _words_pt(item["desc"])
        matches_code = item_code in q_codes if item_code else False
        matches_name = bool(q_words & desc_words)
        if matches_code or (not q_codes and matches_name):
            seen.add(item["desc"])
            matches.append((bucket_name, item))

    return matches[:max_matches]


def _fmt_produto_block(matches: list[tuple[str, dict]]) -> str:
    """Serializa os dados completos de um ou poucos produtos para resposta focada."""
    lines = []
    for bucket, item in matches:
        comprado = (
            f"{item['qtd_comprada']:g} {item['un']} ({item.get('data_compra') or 'sem data'})"
            if item.get("qtd_comprada") else "sem registro de compra"
        )
        lines.append(f"*{item['desc']}*")
        lines.append(f"Classificação: {_BUCKET_LABELS.get(bucket, bucket)}")
        lines.append(
            f"Saldo atual: {item['saldo_atual']:g} {item['un']}  |  "
            f"Comprado para a semana: {comprado}"
        )
        if item.get("motivo"):
            lines.append(f"Motivo: {item['motivo']}")
        lines.append("")
    return "\n".join(lines).strip()


async def hortifruti_report_node(state: State) -> dict:
    """Relatório ou lista de reposição hortifruti, dependendo do intent."""
    cnpj   = state.get("cnpj") or ""
    intent = state.get("intent") or ""
    if intent.startswith("reposicao_mais_"):
        mode = "reposicao_mais"
    elif intent.startswith("reposicao_"):
        mode = "reposicao"
    else:
        mode = "relatorio"
    _log("hortifruti_report", cnpj=cnpj, mode=mode)

    # ── 1. Estoque (últimos 3 meses, granularidade semanal) ───────────────────
    question_for_filter = state.get("question") or ""
    section_requested = bool({"hortifruti", "flv", "lfv", "vlf"} & _words_pt(question_for_filter))
    produto_like_terms = (
        _termos_produto_para_like(question_for_filter)
        if (
            mode in ("reposicao", "reposicao_mais")
            and intent != "reposicao_resumo_hortifruti"
            and not section_requested
        ) else []
    )
    product_filter = "".join(
        f"\n  AND upper(s.PRODUTO_CHAVE) LIKE '%{term}%'"
        for term in produto_like_terms
    )
    eligibility_filter = (
        ""
        if produto_like_terms else """
  AND c.DATA_ULTIMA_COMPRA >= ref.max_data - INTERVAL 21 DAYS
  AND GREATEST(
    COALESCE(v.CONSUMO_CICLO, 0),
    COALESCE(mv.MEDIA_VENDA_SEMANAL, 0),
    COALESCE(h.MEDIA_COMPRA_SEMANAL, 0)
  ) > s.SALDO_ATUAL"""
    )
    sql = cleanup_sql(_HORTIFRUTI_REPORT_SQL.format(
        cnpj=cnpj,
        eligibility_filter=eligibility_filter,
        product_filter=product_filter,
    ))
    try:
        rows = parse_rows(await run_query(sql))
        _log("hortifruti_report", result=f"{len(rows)} registros semanais")
    except Exception as e:
        _log("hortifruti_report", result=f"ERRO estoque: {e}")
        return {"answer": "Não consegui buscar os dados de estoque do hortifruti. Tente novamente em instantes.", "insight": None, "excel_url": None}

    if not rows:
        return {"answer": "Não encontrei movimentações de hortifruti nos últimos 7 dias para o seu estabelecimento.", "insight": None, "excel_url": None}

    triage = _triage_hortifruti(rows, include_all_products=bool(produto_like_terms))
    _log("hortifruti_report",
         total=triage["total"], critico=triage["n_critico"],
         comprar_mais=triage["n_comprar_mais"])

    if intent == "reposicao_resumo_hortifruti":
        export_version = int(datetime.now().timestamp())
        excel_url = (
            f"https://imaisapi.martins.com.br/api/reposicao-hortifruti/export"
            f"?cnpj={cnpj}&v={export_version}"
        )
        answer = "📊 Preparei o resumo da sugestao de compra por fornecedor em uma planilha."
        return {
            "answer": answer,
            "excel_url": excel_url,
            "insight": None,
            "last_question": state.get("question") or "",
            "last_answer": answer,
        }

    # ── Mês atual (usado em ambos os modos) ──────────────────────────────────
    from datetime import date as _date
    _PT_MONTHS = {
        1: "janeiro", 2: "fevereiro", 3: "março",    4: "abril",
        5: "maio",    6: "junho",     7: "julho",     8: "agosto",
        9: "setembro",10: "outubro",  11: "novembro", 12: "dezembro",
    }
    _today = _date.today()
    current_month = f"{_today.day} de {_PT_MONTHS[_today.month]} de {_today.year}"

    # ── Pergunta sobre produto(s) específico(s)? ──────────────────────────────
    # Ex: "sugestão de compra de almeirão" — a lista de reposição deve conter
    # somente o item solicitado, sem itens adicionais.
    question = state.get("question") or ""
    produto_matches = _match_produtos_triage(question, triage)
    if produto_matches and mode == "relatorio":
        _log("hortifruti_report", produto_match=[m[1]["desc"] for m in produto_matches])
        produto_block = _fmt_produto_block(produto_matches)
        try:
            answer = await _generator.generate_hortifruti_produto(
                question=question, data_block=produto_block, current_month=current_month,
            )
        except Exception as e:
            _log("hortifruti_report", result=f"ERRO LLM produto: {e}")
            answer = "Encontrei o produto nos seus dados, mas não consegui montar a resposta agora. Tente novamente."
        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── Clima e "produtos parados" — temporariamente indisponíveis ───────────
    # Dependiam da tendência de consumo semanal (3 meses de histórico), que o novo
    # modelo diário (ENTRADA vs VENDA) não calcula — só saldo atual + última compra.
    # Responder com esses dados degradados daria recomendação errada; melhor admitir
    # que está indisponível por ora do que arriscar um conselho incorreto.
    if _eh_pergunta_clima(question) or _eh_pergunta_parados(question):
        answer = (
            "Essa funcionalidade está temporariamente indisponível enquanto atualizamos "
            "os dados de estoque. Mas posso te dizer quais produtos precisam de reposição "
            "urgente agora — quer que eu mostre?"
        )
        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── Pergunta sobre clima/temperatura para decisão de compra de FLV? ─────────
    # Ex: "como está o tempo para compras de FLV?", "o que comprar hoje no
    # hortifruti de acordo com a temperatura?" — busca clima via wttr.in e
    # cruza com estoque atual para recomendar o que priorizar.
    if _eh_pergunta_clima(question):
        # Cidade do cliente (mesma SQL usada no modo relatorio)
        municipio_clima = ""
        try:
            city_rows_c = parse_rows(await run_query(_HORTIFRUTI_CITY_SQL.format(cnpj=cnpj)))
            municipio_clima = str((city_rows_c or [[""]])[0][0] or "").strip()
        except Exception as e:
            _log("hortifruti_report", result=f"AVISO cidade (clima): {e}")

        _log("hortifruti_report", clima_cidade=municipio_clima or "(não encontrado)")

        clima = None
        if municipio_clima:
            clima = await _buscar_clima_cidade(municipio_clima)
            _log("hortifruti_report", clima=clima)

        estoque_block = _fmt_clima_estoque_block(triage)
        try:
            answer = await _generator.generate_hortifruti_clima(
                question=question,
                cidade=municipio_clima,
                clima=clima,
                estoque_block=estoque_block,
                current_month=current_month,
            )
        except Exception as e:
            _log("hortifruti_report", result=f"ERRO LLM clima: {e}")
            answer = "Encontrei os dados de estoque, mas não consegui buscar o clima agora. Tente de novo."

        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── "Me mostre mais" continuando uma lista de PARADOS (não de reposição)? ──
    # Ex: pergunta "quais estão parados" (mostra top 15 de N) seguida de um
    # follow-up genérico "me mostre mais" — o classificador rotula isso como
    # reposicao_mais_hortifruti (única intenção de "mais" que existe hoje), mas
    # o cliente está pedindo para continuar a lista de PARADOS, não a de reposição.
    # Detectamos pelo assunto da pergunta/resposta anteriores, não pela atual.
    prev_question = state.get("last_question") or ""
    prev_answer   = state.get("last_answer") or ""
    if (
        _eh_continuacao_generica(question)
        and not _eh_pergunta_parados(question)
        and (
            _eh_pergunta_parados(prev_question)
            or prev_answer.startswith("🛑 *Produtos parados")
            or prev_answer.startswith("🛑 *Mais produtos parados")
        )
    ):
        MAX_PARADOS_RESPOSTA = 15
        parados_com_estoque = sorted(
            (i for i in (triage.get("_all_parados") or []) if i["qnt_atu"] > 0),
            key=lambda x: -x["qnt_atu"],
        )
        # pula os que já apareceram na resposta anterior — assim "mostre mais"
        # sempre revela itens novos, nunca repete a mesma página
        restantes = [i for i in parados_com_estoque if i["desc"] not in prev_answer]
        _log("hortifruti_report", parados_continuacao=len(restantes))
        if not restantes:
            answer = "Esses já eram todos os produtos parados com estoque que encontrei — não há mais para mostrar."
        else:
            total = len(restantes)
            mostrar = restantes[:MAX_PARADOS_RESPOSTA]
            linhas = [f"🛑 *Mais produtos parados no hortifruti ({total} restante(s)):*", ""]
            linhas += [f"*{item['desc']}* — Estoque parado: {item['qnt_atu']} {item['un']}" for item in mostrar]
            if total > MAX_PARADOS_RESPOSTA:
                linhas.append("")
                linhas.append(f"...e mais {total - MAX_PARADOS_RESPOSTA} produto(s). Me pergunte para ver mais.")
            answer = "\n".join(linhas)
        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── Pergunta sobre produtos parados (sem giro)? ───────────────────────────
    # Ex: "quais produtos estão parados no hortifruti" — responder direto com a
    # classificação 'parado' já calculada (série semanal de 3 meses), sem LLM,
    # para nunca contradizer os números reais (ex: dizer "sem vendas" sobre um
    # produto que na verdade vendeu centenas de unidades).
    if _eh_pergunta_parados(question):
        MAX_PARADOS_RESPOSTA = 15
        # Só interessam os que TÊM estoque parado (dinheiro parado na prateleira) —
        # itens com estoque zerado são apenas SKUs do catálogo sem giro nenhum,
        # não geram ação nenhuma para o lojista.
        parados_com_estoque = sorted(
            (i for i in (triage.get("_all_parados") or []) if i["qnt_atu"] > 0),
            key=lambda x: -x["qnt_atu"],
        )
        _log("hortifruti_report", parados_count=len(parados_com_estoque))
        if not parados_com_estoque:
            answer = "Nenhum produto do seu hortifruti está parado com estoque sobrando — ótimo sinal! 🎉"
        else:
            total = len(parados_com_estoque)
            mostrar = parados_com_estoque[:MAX_PARADOS_RESPOSTA]
            linhas = [
                f"🛑 *Produtos parados no hortifruti — sem saída nos últimos 30 dias, ainda com estoque ({total}):*",
                "",
            ]
            linhas += [f"*{item['desc']}* — Estoque parado: {item['qnt_atu']} {item['un']}" for item in mostrar]
            if total > MAX_PARADOS_RESPOSTA:
                linhas.append("")
                linhas.append(f"...e mais {total - MAX_PARADOS_RESPOSTA} produto(s) parados com estoque. Me pergunte para ver mais.")
            answer = "\n".join(linhas)
        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── Follow-up genérico após resposta de CLIMA → continua no fluxo de clima ──
    # Ex: "e para o hortifruti" após "qual sugestão para FLV de acordo com o clima?" —
    # o classificador roteia como reposicao_hortifruti (não viu contexto de clima),
    # mas o usuário quer a mesma lógica de temperatura aplicada ao hortifruti.
    # Detectamos pelo assunto da pergunta anterior: se era sobre clima e a pergunta
    # atual é curta e genérica (sem palavras de reposição explícitas), continuamos clima.
    _REPOSICAO_EXPLICITA = {"repor", "reposicao", "lista", "comprar", "compra", "falta", "preciso", "faltando"}
    if (
        mode == "reposicao"
        and _eh_pergunta_clima(prev_question)
        and not _eh_pergunta_clima(question)
        and len(_words_pt(question)) <= 6
        and not (_words_pt(question) & _REPOSICAO_EXPLICITA)
    ):
        # Clima está temporariamente indisponível (ver bloco acima) — não redireciona
        # para o fluxo degradado de clima, só avisa.
        _log("hortifruti_report", result="follow-up clima → indisponível")
        answer = (
            "Essa funcionalidade está temporariamente indisponível enquanto atualizamos "
            "os dados de estoque."
        )
        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": question,
            "last_answer":   answer,
        }

    # ── Modo REPOSIÇÃO (normal ou lista completa) ─────────────────────────────
    if mode in ("reposicao", "reposicao_mais"):
        fornecedores = _fornecedores_solicitados(question, triage)
        if fornecedores:
            sugestoes = _sugestoes_ordenadas_por_giro(triage)
            mensagens_fornecedor: list[str] = []
            for fornecedor in fornecedores:
                itens = [
                    item for item in sugestoes
                    if fornecedor in _fornecedores_do_item(item)
                ]
                if itens:
                    mensagens_fornecedor.append(_fmt_reposicao_block(
                        triage,
                        display_items=itens,
                        fornecedor_filter=fornecedor,
                        title=f"🛒 sugestao de compra — {fornecedor}",
                    ))

            if mensagens_fornecedor:
                answer = (
                    mensagens_fornecedor[0]
                    if len(mensagens_fornecedor) == 1
                    else f"Separei a sugestao de compra em {len(mensagens_fornecedor)} mensagens, uma por fornecedor."
                )
                return {
                    "answer": answer,
                    "output_messages": mensagens_fornecedor,
                    "output_messages_question": question,
                    "insight": None,
                    "excel_url": None,
                    "last_question": question,
                    "last_answer": "🛒 sugestao de compra por fornecedor",
                }
        if produto_matches:
            descs = {item["desc"] for _, item in produto_matches}
            _log("hortifruti_report", reposicao_produto=list(descs))
            codes = _product_codes_in_question(question)
            answer = _fmt_reposicao_block(
                triage,
                only_codes=codes or None,
                only_descs=None if codes else descs,
            )
        elif _product_codes_in_question(question):
            answer = "Não encontrei um produto com esse codigo na sugestao de compra."
        else:
            sugestoes = _sugestoes_ordenadas_por_giro(triage)
            if mode == "reposicao_mais":
                answer = "A sugestao de compra completa já foi exibida."
            else:
                answer = _fmt_reposicao_block(
                    triage,
                    display_items=sugestoes,
                    title="🛒 sugestao de compra — Hortifruti",
                )

        return {
            "answer":        answer,
            "insight":       None,
            "excel_url":     None,
            "last_question": state.get("question") or "",
            "last_answer":   answer,
        }

    # ── Modo RELATÓRIO: relatório completo com mercado e sazonalidade ─────────
    data_block = _fmt_triage_block(triage)

    # Cidade do cliente
    municipio = ""
    try:
        city_rows = parse_rows(await run_query(_HORTIFRUTI_CITY_SQL.format(cnpj=cnpj)))
        municipio = str((city_rows or [[""]])[0][0] or "").strip()
        _log("hortifruti_report", municipio=municipio or "(não encontrado)")
    except Exception as e:
        _log("hortifruti_report", result=f"AVISO cidade: {e}")

    # Top produtos do mercado local
    market_block = ""
    if municipio:
        try:
            import unicodedata as _ud
            def _norm(s: str) -> str:
                s = s.lower().strip()
                return "".join(c for c in _ud.normalize("NFD", s) if _ud.category(c) != "Mn")

            mkt_rows     = parse_rows(await run_query(_HORTIFRUTI_MARKET_SQL.format(municipio=municipio)))
            client_names = {_norm(str(r[1])) for r in rows if r and r[1]}
            opportunities: list[str] = []
            for mr in (mkt_rows or []):
                if not mr or not mr[0]:
                    continue
                prod  = str(mr[0])
                qtd   = mr[1] if len(mr) > 1 else 0
                lojas = mr[2] if len(mr) > 2 else 0
                words = [w for w in _norm(prod).split() if len(w) >= 4]
                if not any(any(w in cn for cn in client_names) for w in words):
                    opportunities.append(f"  {prod} | vendido por {lojas} loja(s) | {qtd} unid/90d")
                if len(opportunities) >= 15:
                    break
            if opportunities:
                market_block = (
                    f"Cidade: {municipio}\n"
                    "Produtos hortifruti mais vendidos na região que você NÃO tem no estoque:\n"
                    + "\n".join(opportunities)
                )
                _log("hortifruti_report", oportunidades=len(opportunities))
        except Exception as e:
            _log("hortifruti_report", result=f"AVISO mercado: {e}")

    try:
        answer = await _generator.generate_hortifruti_report(
            data_block=data_block, market_block=market_block,
            municipio=municipio, current_month=current_month,
        )
    except Exception as e:
        _log("hortifruti_report", result=f"ERRO LLM relatorio: {e}")
        answer = "Dados encontrados, mas não consegui gerar o relatório. Tente novamente."

    return {
        "answer":        answer,
        "insight":       None,
        "excel_url":     None,
        "last_question": state.get("question") or "",
        "last_answer":   answer,
    }


async def perdas_report_node(state: State) -> dict:
    """Histórico de movimentações manuais do estoque (silver_prod.estoque_quantum_poc)."""
    cnpj     = state.get("cnpj") or ""
    question = state.get("question") or ""
    _log("perdas_report", cnpj=cnpj)

    data_filter, periodo, data_inicio, data_fim = _parse_perdas_data_filter(question)

    sql = cleanup_sql(_PERDAS_HISTORICO_SQL.format(cnpj=cnpj, data_filter=data_filter))
    try:
        rows = parse_rows(await run_query(sql))
        _log("perdas_report", result=f"{len(rows)} movimentações", periodo=periodo)
    except Exception as e:
        _log("perdas_report", result=f"ERRO: {e}")
        return {"answer": "Não consegui buscar o histórico de movimentações. Tente novamente em instantes.", "insight": None}

    excel_url = (
        f"https://imaisapi.martins.com.br/api/perdas/export"
        f"?cnpj={cnpj}&data_inicio={data_inicio}&data_fim={data_fim}"
    )
    answer = _fmt_perdas_block(rows, periodo=periodo)
    return {
        "answer":        answer,
        "excel_url":     excel_url,
        "insight":       None,
        "last_question": question,
        "last_answer":   answer,
    }


async def promocao_node(state: State) -> dict:
    """Sugere produtos para promoção com preço recomendado, cruzando a curva ABCD
    do cliente com os preços da concorrência na mesma região (UF)."""
    cnpj     = state.get("cnpj") or ""
    question = state.get("question") or ""
    _log("promocao", cnpj=cnpj)

    sql = cleanup_sql(_PROMOCAO_SQL.format(cnpj=cnpj))
    try:
        js   = await run_query(sql)
        cols = parse_cols(js)
        rows = parse_rows(js)
        _log("promocao", result=f"{len(rows)} candidatos")
    except Exception as e:
        _log("promocao", result=f"ERRO SQL: {e}")
        return {
            "answer": "Não consegui montar a sugestão de promoção agora. Tente novamente em instantes.",
            "insight": None,
        }

    if not rows:
        answer = (
            "Não encontrei produtos com espaço claro para promoção no momento — seus preços já estão "
            "competitivos frente à concorrência da sua região, ou ainda não tenho comparativo de mercado "
            "suficiente para esses itens. Quando tiver mais dados, te trago boas oportunidades de oferta. 👍"
        )
        return {"answer": answer, "insight": None, "last_question": question, "last_answer": answer}

    # Região (UF) — só para dar contexto na resposta
    regiao = ""
    try:
        # a UF não vem na projeção; busca rápida da UF predominante do cliente
        uf_rows = parse_rows(await run_query(
            f"SELECT UF FROM imaiscatalog.gold_prod.mvp_dados_intermediarios "
            f"WHERE CNPJ = '{cnpj}' AND UF IS NOT NULL AND trim(UF) <> '' "
            f"GROUP BY UF ORDER BY COUNT(*) DESC LIMIT 1"
        ))
        regiao = str((uf_rows or [[""]])[0][0] or "").strip()
    except Exception as e:
        _log("promocao", result=f"AVISO UF: {e}")

    from datetime import date as _date
    _PT_MONTHS = {
        1: "janeiro", 2: "fevereiro", 3: "março",    4: "abril",
        5: "maio",    6: "junho",     7: "julho",     8: "agosto",
        9: "setembro",10: "outubro",  11: "novembro", 12: "dezembro",
    }
    _today = _date.today()
    current_month = f"{_today.day} de {_PT_MONTHS[_today.month]} de {_today.year}"

    data_block = _fmt_promocao_block(rows, cols)
    try:
        answer = await _generator.generate_promocao(
            question=question, data_block=data_block,
            regiao=regiao, current_month=current_month,
        )
    except Exception as e:
        _log("promocao", result=f"ERRO LLM: {e}")
        answer = "Encontrei os produtos, mas não consegui montar a recomendação de preço agora. Tente novamente."

    return {
        "answer":       answer,
        "insight":      None,
        "last_question": question,
        "last_answer":  answer,
    }


_ESTOQUE_TRANSLATE = (
    "translate(lower(DESCRICAO), 'áàâãéèêíìîóòôõúùûüç', 'aaaaeeeiiioooouuuuc')"
)

# Template para estoque_produto: busca produto específico (com filtro de nome/GTIN)
_ESTOQUE_PRODUTO_TEMPLATE = """\
WITH latest_e AS (
  SELECT
    e.CEAN AS GTIN,
    e.CODIGO,
    e.DESCRICAO,
    e.UNIDADE_MEDIDA,
    e.QNT_ESTOQUE,
    ROW_NUMBER() OVER (PARTITION BY e.CODIGO ORDER BY e.DATA_RELATORIO DESC) AS rn
  FROM imaiscatalog.silver_prod.estoque_quantum_poc e
  JOIN imaiscatalog.gold_prod.dim_cli c ON c.SRK_CLI = e.SRK_CLI
  WHERE lpad(cast(c.CNPJ_CPF as string), 14, '0') = '{cnpj}'
)
SELECT GTIN, CODIGO, DESCRICAO, UNIDADE_MEDIDA, QNT_ESTOQUE
FROM latest_e
WHERE rn = 1{filter_clause}
ORDER BY QNT_ESTOQUE DESC
LIMIT 20"""

# Template para estoque_nivel: visão geral do estoque, ordenada por vendas recentes
_ESTOQUE_NIVEL_TEMPLATE = """\
WITH latest_e AS (
  SELECT
    e.CEAN AS GTIN,
    e.CODIGO,
    e.DESCRICAO,
    e.UNIDADE_MEDIDA,
    e.QNT_ESTOQUE,
    ROW_NUMBER() OVER (PARTITION BY e.CODIGO ORDER BY e.DATA_RELATORIO DESC) AS rn
  FROM imaiscatalog.silver_prod.estoque_quantum_poc e
  JOIN imaiscatalog.gold_prod.dim_cli c ON c.SRK_CLI = e.SRK_CLI
  WHERE lpad(cast(c.CNPJ_CPF as string), 14, '0') = '{cnpj}'
),
estoque_atual AS (
  SELECT GTIN, CODIGO, DESCRICAO, UNIDADE_MEDIDA, QNT_ESTOQUE
  FROM latest_e
  WHERE rn = 1{filter_clause}
),
vendas_recentes AS (
  SELECT
    CODIGO_GTIN,
    SUM(CAST(QUANTIDADE_COMPRADA AS DECIMAL(15,2))) AS QTD_VENDIDA_90D
  FROM imaiscatalog.gold_prod.mvp_dados_intermediarios
  WHERE CNPJ = '{cnpj}'
    AND TIPO_NOTA = 'SAIDA'
    AND to_date(DATA_EMISSAO) >= date_sub(current_date(), 90)
  GROUP BY CODIGO_GTIN
)
SELECT
  e.GTIN,
  e.CODIGO,
  e.DESCRICAO,
  e.UNIDADE_MEDIDA,
  e.QNT_ESTOQUE,
  coalesce(v.QTD_VENDIDA_90D, 0) AS QTD_VENDIDA_90D
FROM estoque_atual e
LEFT JOIN vendas_recentes v ON v.CODIGO_GTIN = e.GTIN
ORDER BY QTD_VENDIDA_90D DESC, e.QNT_ESTOQUE DESC
LIMIT 20"""


def _build_estoque_filter(product_filter: str, asked_for_zero: bool) -> str:
    """Converte product_filter em cláusula WHERE para o template de estoque.
    Expande variantes singulares automaticamente (alfaces → alface).
    """
    pf = (product_filter or "").strip()
    parts: list[str] = []

    if pf:
        words = [w.lower() for w in pf.split() if len(w) >= 2]
        if not words:
            words = [pf.lower()]
        for w in words:
            variants = list(dict.fromkeys(_singular_variants(w)))
            or_part = " OR ".join(
                f"{_ESTOQUE_TRANSLATE} LIKE '%{v}%'" for v in variants
            )
            parts.append(f"({or_part})" if len(variants) > 1 else or_part)
        filter_sql = " AND ".join(parts)
        clause = f"\n  AND ({filter_sql})"
    elif asked_for_zero:
        clause = "\n  AND QNT_ESTOQUE <= 0"
    else:
        clause = "\n  AND QNT_ESTOQUE > 0"  # estoque_nivel: só produtos com saldo positivo

    return clause


def _build_estoque_sql(cnpj: str, params_dict: dict | None, question: str) -> str | None:
    """Retorna SQL a partir do template fixo para queries de estoque.
    Usa template com join de vendas para visão geral (estoque_nivel),
    e template simples para produto específico (estoque_produto).
    Retorna None quando não é uma query de estoque (deixa LLM gerar).
    """
    metric     = (params_dict or {}).get("metric", "")
    pref_table = (params_dict or {}).get("preferred_table", "")

    is_estoque = metric in ("estoque_produto", "estoque_nivel") or pref_table == "estoque_quantum_poc"
    if not is_estoque:
        return None

    product_filter = (params_dict or {}).get("product_filter") or ""
    asked_for_zero = any(kw in (question or "").lower() for kw in _ZEROED_KEYWORDS)
    filter_clause  = _build_estoque_filter(product_filter, asked_for_zero)

    # estoque_nivel sem produto → usa template com join de vendas (top sellers)
    if metric == "estoque_nivel" and not product_filter:
        return _ESTOQUE_NIVEL_TEMPLATE.format(cnpj=cnpj, filter_clause=filter_clause)

    # estoque_produto ou visão com filtro de produto → template simples com LIMIT 20
    return _ESTOQUE_PRODUTO_TEMPLATE.format(cnpj=cnpj, filter_clause=filter_clause)


async def sql_gen_node(state: State) -> dict:
    """Gera SQL via LLM e valida antes de executar.
    Para queries de estoque usa template fixo em vez do LLM.
    """
    question       = state.get("question") or ""
    cnpj           = state.get("cnpj") or ""
    attempts       = state.get("sql_attempts") or 0
    error_feedback = state.get("sql_error")
    today          = date.today().isoformat()

    # Reconstrói ExtractedParams do dict salvo no state
    params_dict = state.get("extracted_params")
    extracted = ExtractedParams(**params_dict) if params_dict else None

    attempts += 1
    _log("sql_gen", attempt=attempts, error_feedback=error_feedback or "nenhum")

    # ── Template fixo para estoque (primeira tentativa sem erro prévio) ────────
    if attempts == 1 and not error_feedback:
        template_sql = _build_estoque_sql(cnpj, params_dict, question)
        if template_sql:
            sql = cleanup_sql(template_sql)
            _log("sql_gen", result="SQL via template (estoque) ✓", sql=sql[:80])
            return {
                "sql":          sql,
                "sql_note":     None,
                "sql_attempts": attempts,
                "sql_error":    None,
                "supervisor_retry": False,
            }

    # ── LLM gera SQL (para não-estoque ou retry de estoque com erro) ──────────
    try:
        result = await _generator.generate_sql(
            question=question,
            cnpj=cnpj,
            today=today,
            error_feedback=error_feedback,
            extracted_params=extracted,
        )
        sql = cleanup_sql(result.sql or "")
    except Exception as e:
        _log("sql_gen", result=f"ERRO geração: {e}")
        return {
            "sql":          None,
            "sql_attempts": attempts,
            "sql_error":    f"Erro ao gerar SQL: {e}",
        }

    _log("sql_gen", sql=sql[:120])

    # Valida segurança e tabelas
    ok, reason = validate_sql_against_schema(sql, _schema)
    if not ok:
        _log("sql_gen", result=f"VALIDAÇÃO FALHOU: {reason}")
        return {
            "sql":          None,
            "sql_attempts": attempts,
            "sql_error":    f"SQL inválida ({reason}). Gere uma SQL diferente usando apenas as tabelas do schema.",
        }

    # Preprocessamento: corrige problemas de estoque direto no SQL (sem rejeitar)
    sql = _preprocess_estoque_sql(sql, params_dict, question)

    # Validação de anti-padrões restantes (que precisam de retry com feedback)
    anti = _check_estoque_anti_patterns(sql, params_dict, question)
    if anti and attempts < 3:
        _log("sql_gen", result=f"ANTI-PADRÃO: {anti[:80]}")
        return {
            "sql":          None,
            "sql_attempts": attempts,
            "sql_error":    anti,
        }

    _log("sql_gen", result="SQL válida ✓")
    return {
        "sql":          sql,
        "sql_note":     result.note or None,
        "sql_attempts": attempts,
        "sql_error":    None,
        "supervisor_retry": False,
    }


_SUM_RE             = re.compile(r"\bSUM\s*\(", re.IGNORECASE)
_GROUP_BY_RE        = re.compile(r"\n?\s*GROUP\s+BY\b[^\n]*(?:\n[^\n]+)*", re.IGNORECASE)
_ZEROED_KEYWORDS    = ("zerado", "zero", "ruptura", "sem estoque", "faltando", "em falta", "esgotado")
# Remove cláusula AND QNT_ESTOQUE <= 0 (com ou sem qualificador de tabela)
_QNT_ZERO_STRIP_RE  = re.compile(
    r"\s*AND\s+(?:\w+\.)?QNT_ESTOQUE\s*<=?\s*0\b",
    re.IGNORECASE,
)
_BARE_GTIN_RE       = re.compile(r"\bGTIN\b", re.IGNORECASE)
_CEAN_AS_GTIN_RE    = re.compile(r"\bCEAN\s+AS\s+GTIN\b", re.IGNORECASE)
_GTIN_PLACEHOLDER   = "__CEAN_AS_GTIN__"


def _preprocess_estoque_sql(sql: str, params_dict: dict | None, question: str) -> str:
    """Aplica todas as correções programáticas em queries de estoque.
    Não rejeita nem pede retry — corrige cirurgicamente e segue.
    """
    if not _ESTOQUE_RE.search(sql or ""):
        return sql

    metric     = (params_dict or {}).get("metric", "")
    asked_zero = any(kw in (question or "").lower() for kw in _ZEROED_KEYWORDS)
    fixes: list[str] = []

    # 1. Remove QNT_ESTOQUE <= 0 quando não foi solicitado
    if not asked_zero and _QNT_ZERO_RE.search(sql):
        sql = _QNT_ZERO_STRIP_RE.sub("", sql)
        fixes.append("removido filtro QNT_ESTOQUE<=0")

    # 2. Para estoque_produto: remove GROUP BY sem SUM (ROW_NUMBER já deduplica)
    if metric == "estoque_produto" and _GROUP_BY_RE.search(sql) and not _SUM_RE.search(sql):
        sql = _GROUP_BY_RE.sub("", sql)
        fixes.append("removido GROUP BY sem SUM (ROW_NUMBER já garante unicidade)")

    # 3. Corrige `e.CEAN AS GTIN` dentro da CTE (alias na posição errada).
    #    Dentro da CTE, CEAN deve ser selecionado sem alias (e.CEAN).
    #    O alias AS GTIN pertence ao SELECT externo, não à CTE.
    #    `e.CEAN` com prefixo de tabela só aparece dentro da CTE.
    _cte_alias_re = re.compile(r'\be\.CEAN\s+AS\s+GTIN\b', re.IGNORECASE)
    if _cte_alias_re.search(sql):
        sql = _cte_alias_re.sub('e.CEAN', sql)
        fixes.append("movido alias GTIN da CTE para o SELECT externo")

    # 4. Corrige GTIN sem alias no SELECT externo: bare GTIN → CEAN AS GTIN.
    #    Dois passos para não duplicar aliases já corretos.
    if _BARE_GTIN_RE.search(sql):
        sql = _CEAN_AS_GTIN_RE.sub(_GTIN_PLACEHOLDER, sql)      # protege já corretos
        sql = _BARE_GTIN_RE.sub("CEAN AS GTIN", sql)             # corrige bare GTIN
        sql = sql.replace(_GTIN_PLACEHOLDER, "CEAN AS GTIN")     # restaura
        fixes.append("corrigido GTIN → CEAN AS GTIN")

    if fixes:
        print(f"[sql_gen] Preprocessado estoque: {' | '.join(fixes)}")
    return sql


def _check_estoque_anti_patterns(sql: str, params_dict: dict | None, question: str) -> str | None:
    """Detecta anti-padrões que NÃO podem ser corrigidos por preprocessamento.
    Retorna feedback para retry, ou None se tudo OK.
    """
    if not _ESTOQUE_RE.search(sql or ""):
        return None
    metric = (params_dict or {}).get("metric", "")

    # SUM em estoque_produto requer GROUP BY e distorce a listagem individual
    if metric == "estoque_produto" and _SUM_RE.search(sql or ""):
        return (
            "ANTI-PADRÃO: estoque_produto exige LINHAS INDIVIDUAIS por produto, "
            "mas a SQL usa SUM() (agregação). "
            "Gere SELECT simples: CEAN AS GTIN, CODIGO, DESCRICAO, UNIDADE_MEDIDA, QNT_ESTOQUE — sem SUM, sem GROUP BY."
        )
    return None


async def execute_node(state: State) -> dict:
    """Executa a SQL no Databricks."""
    sql = state.get("sql") or ""
    _log("execute", sql=sql[:120])

    try:
        js      = await run_query(sql)
        columns = parse_cols(js)
        rows    = parse_rows(js)
        _log("execute", result=f"{len(rows)} rows, cols={columns}")
        return {"columns": columns, "rows": rows}
    except Exception as e:
        _log("execute", result=f"ERRO: {e}")
        return {
            "columns": [],
            "rows":    [],
            "sql_error": f"Erro ao executar SQL no Databricks: {e}",
            "supervisor_retry": False,
        }


# Colunas hierárquicas de classificação de produto — quando um filtro em uma delas
# retorna vazio, o supervisor sugere testar as demais.
_HIER_COLS = ["CATEGORIA", "SECAO", "SUBCATEGORIA", "DEPARTAMENTO", "MARCA", "FABRICANTE"]
_HIER_ILIKE_RE = re.compile(
    r"\b(" + "|".join(_HIER_COLS) + r")\s+ILIKE\s+'%([^']+)%'",
    re.IGNORECASE,
)
_PROD_ILIKE_RE = re.compile(
    r"(?:\bDESC_PROD\s+ILIKE\s+'%([^']+)%'"                           # DESC_PROD ILIKE '%term%'
    r"|translate\(lower\(i\.DESC_PROD\)[^)]*\)\s+LIKE\s+'%([^']+)%'"  # translate(lower(i.DESC_PROD),...) LIKE '%term%'
    r")",
    re.IGNORECASE,
)


def _rows_are_empty(rows: list, columns: list | None = None) -> bool:
    """Retorna True se rows está vazio ou se todos os valores analíticos são None/nulos.
    Ignora colunas de contagem (n_*, qtd_*, count_*) para não tratar '0 pares' como vazio.
    """
    if not rows:
        return True
    # Identifica índices de colunas que são apenas contadores — não são dados analíticos
    count_cols: set[int] = set()
    if columns:
        for i, col in enumerate(columns):
            c = str(col).lower()
            if c.startswith(("n_", "qtd_", "count_", "quantidade_", "n_pares", "n_meses")):
                count_cols.add(i)
    for row in rows:
        for i, val in enumerate(row or []):
            if i in count_cols:
                continue
            if val is not None and str(val).strip() not in ("", "0", "0.0", "0.00"):
                return False
    return True


_INTERMEDIARIOS_RE      = re.compile(r"\bmvp_dados_intermediarios\b", re.IGNORECASE)
_ESTOQUE_RE             = re.compile(r"\bestoque_quantum_poc\b", re.IGNORECASE)
_DBX_ERROR_CODE_RE      = re.compile(r'\[([A-Z_]+(?:\.[A-Z_]+)*)\]')
_DBX_DOUBLE_ALIAS_RE    = re.compile(r'\bAS\s+\S+\s+AS\s+\S+', re.IGNORECASE)


def _interpret_sql_error(error_msg: str, sql: str = "") -> str:
    """Converte erros brutos do Databricks em feedback acionável para o LLM."""
    code_match = _DBX_ERROR_CODE_RE.search(error_msg or "")
    code = code_match.group(1) if code_match else ""

    # ── UNRESOLVED_COLUMN: coluna não existe ──────────────────────────────────
    if "UNRESOLVED_COLUMN" in code:
        col_match = re.search(r"name `([^`]+)`", error_msg)
        col = col_match.group(1) if col_match else re.search(r"'(\w+)' cannot be resolved", error_msg, re.I)
        col = col.group(1) if hasattr(col, "group") else str(col)
        suggs = re.findall(r'`([A-Z_0-9]+)`', error_msg)
        suggs = [s for s in suggs if s != col]

        base = f"Coluna `{col}` não existe nesse contexto."
        if suggs:
            base += f" Colunas disponíveis: {', '.join(suggs[:6])}."

        # Casos específicos GTIN/CEAN
        if col.upper() == "GTIN" and any(s.upper() == "CEAN" for s in suggs):
            return (
                f"{base} "
                "A CTE não alias CEAN. No SELECT externo, escreva `CEAN AS GTIN` (não apenas `GTIN`)."
            )
        if col.upper() == "CEAN" and any(s.upper() == "GTIN" for s in suggs):
            return (
                f"{base} "
                "A CTE aliasou CEAN como GTIN. No SELECT externo, use `GTIN` (não `CEAN`)."
            )
        return base + " Corrija o nome da coluna."

    # ── PARSE_SYNTAX_ERROR: erro de sintaxe ──────────────────────────────────
    if "PARSE_SYNTAX_ERROR" in code:
        near_match = re.search(r"at or near '(\w+)'", error_msg)
        near = near_match.group(1) if near_match else "?"
        # Alias duplo (ex: CAST(x) AS col AS alias)
        if _DBX_DOUBLE_ALIAS_RE.search(sql):
            double = _DBX_DOUBLE_ALIAS_RE.search(sql)
            return (
                f"Alias duplo inválido: `{double.group(0)}`. "
                "SQL não aceita dois AS em sequência. "
                "Para renomear: use apenas `coluna AS nome` — sem CAST desnecessário. "
                "Para CEAN: `CEAN AS GTIN` (sem CAST)."
            )
        return (
            f"Erro de sintaxe próximo de `{near}`. "
            "Revise a cláusula nesse ponto. Verifique parênteses, vírgulas e aliases."
        )

    # ── WRONG_NUM_ARGS: função com argumentos errados ─────────────────────────
    if "WRONG_NUM_ARGS" in code:
        m = re.search(r"`(\w+)` requires (\d+) parameters? but .* (\d+)", error_msg, re.I)
        if m:
            func, needed, got = m.group(1), m.group(2), m.group(3)
            tips = {
                "trunc":      "Use: trunc(expr, 'MM') para início do mês, trunc(expr, 'YYYY') para ano.",
                "date_trunc": "Use: date_trunc('MONTH', expr).",
                "lpad":       "Use: lpad(cast(col as string), tamanho, '0').",
            }
            tip = tips.get(func.lower(), "Verifique a documentação do Databricks Spark SQL.")
            return f"Função `{func}` requer {needed} argumento(s), recebeu {got}. {tip}"

    # ── UNRESOLVED_ROUTINE / UNDEFINED_FUNCTION ───────────────────────────────
    if "UNRESOLVED_ROUTINE" in code or "UNDEFINED_FUNCTION" in code:
        m = re.search(r"[Ff]unction\s+`?(\w+)`?", error_msg)
        func = m.group(1) if m else "?"
        fixes = {
            "date_from_parts": "Não existe no Databricks. Use to_date(concat(year,'-',month,'-',day)).",
            "ilike":           "Não existe no Databricks. Use translate()+LIKE para case-insensitive.",
            "ifnull":          "Use coalesce(expr, default) em vez de ifnull().",
            "nvl":             "Use coalesce(expr, default).",
            "to_timestamp_ltz":"Use to_timestamp() em Databricks.",
        }
        tip = fixes.get(func.lower(), "Verifique se a função existe no Databricks Spark SQL.")
        return f"Função `{func}` não existe no Databricks. {tip}"

    # ── AMBIGUOUS_REFERENCE: coluna ambígua ───────────────────────────────────
    if "AMBIGUOUS_REFERENCE" in code:
        m = re.search(r"`(\w+)`", error_msg)
        col = m.group(1) if m else "?"
        return (
            f"Coluna `{col}` é ambígua — existe em mais de uma tabela. "
            f"Qualifique com o alias da tabela: ex. `tabela.{col}`."
        )

    # ── ANALYSIS_ERROR / genérico ─────────────────────────────────────────────
    if code:
        return f"[{code}] {error_msg.split(chr(10))[0].strip()}"

    return error_msg
_DESCRICAO_LIKE_RE  = re.compile(
    r"translate\(lower\((?:e\.)?DESCRICAO\)[^)]*\)\s+LIKE\s+'%([^']+)%'",
    re.IGNORECASE,
)
_QNT_ZERO_RE = re.compile(r"QNT_ESTOQUE\s*<=?\s*0", re.IGNORECASE)


def _singular_variants(word: str) -> list[str]:
    """Gera variantes singulares de uma palavra portuguesa para fallback de busca.
    Sempre inclui a palavra original mais variantes prováveis no singular.
    """
    w = (word or "").lower().strip()
    if not w:
        return []
    seen = {w}
    out = [w]
    def _add(v: str):
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    # ões → ão (questões → questão), ães → ão (pães → pão)
    if len(w) > 4 and (w.endswith("oes") or w.endswith("aes")):
        _add(w[:-3] + "ao")
    # is → l (papéis → papel; após translate vira papeis)
    elif len(w) > 4 and w.endswith("eis"):
        _add(w[:-3] + "el")
    # es → drop (alfaces → alface, simples → simple)
    elif len(w) > 4 and w.endswith("es"):
        _add(w[:-2])
        _add(w[:-1])  # também tenta drop só do 's'
    # s → drop (camisas → camisa)
    elif len(w) > 3 and w.endswith("s"):
        _add(w[:-1])
    return out


_TRANSLATE_EXPR = "translate(lower(i.{col}), 'áàâãéèêíìîóòôõúùûüç', 'aaaaeeeiiioooouuuuc')"


def _build_hier_feedback(sql: str, params_dict: dict | None = None) -> str:
    """Detecta o motivo do resultado vazio e monta feedback direcionado."""
    metric = (params_dict or {}).get("metric", "")

    # Estoque: resultado vazio → ampliar filtro de DESCRICAO, mas manter linhas individuais
    if _ESTOQUE_RE.search(sql or ""):
        desc_match = _DESCRICAO_LIKE_RE.search(sql or "")
        if desc_match:
            term = desc_match.group(1)
            words = [w.lower() for w in term.split() if len(w) >= 3]
            # Expande cada palavra com variantes singulares (alfaces → alface, etc.)
            all_variants: list[str] = []
            seen_v: set[str] = set()
            for w in words:
                for v in _singular_variants(w):
                    if v not in seen_v:
                        seen_v.add(v)
                        all_variants.append(v)
            if all_variants:
                filters = " OR ".join(
                    f"translate(lower(DESCRICAO), 'áàâãéèêíìîóòôõúùûüç','aaaaeeeiiioooouuuuc') LIKE '%{v}%'"
                    for v in all_variants
                )
                return (
                    f"Nenhum produto com '{term}' encontrado no estoque (provavelmente o termo está no plural). "
                    f"Use OBRIGATORIAMENTE este filtro com variantes singulares: ({filters}). "
                    f"IMPORTANTE: mantenha SELECT com linhas individuais por produto (CEAN, CODIGO, DESCRICAO, UNIDADE_MEDIDA, QNT_ESTOQUE). "
                    f"NUNCA use SUM nem GROUP BY. NUNCA filtre QNT_ESTOQUE <= 0 (isso retorna apenas zerados)."
                )
        return (
            "Nenhum produto encontrado no estoque. Tente um filtro de nome mais genérico (forma singular, 1 palavra). "
            "IMPORTANTE: mantenha linhas individuais por produto (SELECT sem SUM). "
            "NUNCA filtre QNT_ESTOQUE <= 0."
        )

    # Se a métrica é comparacao_mercado mas a SQL usou mvp_dados_intermediarios
    if metric == "comparacao_mercado" and _INTERMEDIARIOS_RE.search(sql or ""):
        return (
            "A query usou mvp_dados_intermediarios mas para comparacao_mercado sem produto específico "
            "você DEVE usar nova_mvp_vendas. "
            "Gere uma query com: SELECT v.VALOR_CLIENTE, v.VALOR_MEDIO_MERCADO_MENSAL, "
            "v.TICKET_MEDIO_CLIENTE, v.TICKET_MEDIO_MERCADO, v.NOTAS_CLIENTE, v.LOJAS_CONCORRENTES "
            "FROM imaiscatalog.gold_prod.nova_mvp_vendas v "
            "WHERE v.CNPJ = '<cnpj>' AND v.ANO_MES = date_format(current_date(), 'yyyy-MM') LIMIT 1"
        )

    # Produto não encontrado via DESC_PROD → tenta palavras individuais + colunas hierárquicas
    prod_match = _PROD_ILIKE_RE.search(sql or "")
    if prod_match:
        term = prod_match.group(1) or prod_match.group(2)  # group 1 = ILIKE, group 2 = translate LIKE
        words = [w.lower() for w in term.split() if len(w) >= 3 and w.upper() not in ("AND", "OR", "NOT", "DE", "DA", "DO", "DAS", "DOS")]
        hier_cols = ["CATEGORIA", "SECAO", "SUBCATEGORIA", "DEPARTAMENTO"]
        if words:
            prod_parts = " OR ".join(f"{_TRANSLATE_EXPR.format(col='DESC_PROD')} LIKE '%{w}%'" for w in words)
            hier_parts = " OR ".join(f"{_TRANSLATE_EXPR.format(col=c)} LIKE '%{words[0]}%'" for c in hier_cols)
            full_filter = f"({prod_parts} OR {hier_parts})"
        else:
            full_filter = f"({_TRANSLATE_EXPR.format(col='DESC_PROD')} LIKE '%{term.lower()}%')"
        return (
            f"O produto '{term}' não retornou resultados. "
            f"Na próxima tentativa, amplie a busca usando cada palavra separadamente E buscando "
            f"também nas colunas de categoria, SEMPRE com translate() para acentos. "
            f"Use OBRIGATORIAMENTE este filtro: {full_filter}."
        )

    matches = _HIER_ILIKE_RE.findall(sql or "")
    if not matches:
        return (
            "A query retornou zero resultados. Verifique se os filtros estão corretos, "
            "tente filtros mais abrangentes ou amplie o período de datas."
        )

    used_cols  = [m[0].upper() for m in matches]
    used_terms = list({m[1] for m in matches})
    all_cols   = ["CATEGORIA", "SECAO", "SUBCATEGORIA", "DEPARTAMENTO"]

    term = used_terms[0] if used_terms else "TERMO"
    or_clause = " OR ".join(
        f"{_TRANSLATE_EXPR.format(col=c)} LIKE '%{term.lower()}%'" for c in all_cols
    )

    return (
        f"O filtro '{term}' aplicado apenas em {', '.join(used_cols)} retornou zero resultados. "
        f"O termo pode estar em outra coluna hierárquica. "
        f"Na próxima tentativa, use OBRIGATORIAMENTE todas as colunas com OR e translate() para acentos: "
        f"({or_clause}). "
        f"Copie este filtro exatamente. Não use apenas uma coluna."
    )


async def supervisor_node(state: State) -> dict:
    """Verifica se o resultado é plausível. Pede retry se necessário ou anota
    problemas de qualidade de dados para o writer tratar sem novo retry.
    """
    rows        = state.get("rows") or []
    columns     = state.get("columns") or []
    sql_err     = state.get("sql_error")
    sql         = state.get("sql") or ""
    attempts    = state.get("sql_attempts") or 0
    params_dict = state.get("extracted_params") or {}
    question    = state.get("question") or ""
    metric      = params_dict.get("metric", "")
    grain       = params_dict.get("grain", "")
    limit       = params_dict.get("limit") or 10
    col_names   = [str(c).upper() for c in columns]
    quality_notes: list[str] = []

    _log("supervisor", rows=len(rows), sql_error=bool(sql_err), attempts=attempts)

    # ── Fase 1: Erros e retries ───────────────────────────────────────────────

    # Erro de execução → interpreta e faz retry se ainda tiver tentativas
    if sql_err and attempts < 3:
        interpreted = _interpret_sql_error(sql_err, sql)
        _log("supervisor", result=f"RETRY (erro SQL) | {interpreted[:80]}")
        return {"supervisor_retry": True, "sql": None, "sql_error": interpreted}

    # Resultado vazio → retry com feedback inteligente
    if _rows_are_empty(rows, columns) and attempts < 3:
        feedback = _build_hier_feedback(sql, params_dict)
        _log("supervisor", result=f"RETRY (vazio) | feedback={feedback[:80]}")
        return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # Poucos resultados para lista de produtos → tenta fallback hierárquico
    if grain == "produto" and limit >= 5 and 0 < len(rows) < 3 and attempts < 3 and _HIER_ILIKE_RE.search(sql):
        feedback = _build_hier_feedback(sql, params_dict)
        _log("supervisor", result=f"RETRY (poucos resultados: {len(rows)}) | feedback={feedback[:80]}")
        return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # Produto buscado com frase exata multi-palavra → reescreve como AND-word
    if attempts < 2:
        prod_match = _PROD_ILIKE_RE.search(sql)
        if prod_match:
            term = prod_match.group(1) or prod_match.group(2)
            meaningful_words = [
                w.lower() for w in term.split()
                if len(w) >= 3 and w.upper() not in ("AND", "OR", "NOT", "DE", "DA", "DO", "DAS", "DOS")
            ]
            has_multiword_phrase = len(meaningful_words) >= 2 and " " in term
            if has_multiword_phrase:
                _tr = _TRANSLATE_EXPR.format(col="DESC_PROD")
                and_filter = " AND ".join(f"{_tr} LIKE '%{w}%'" for w in meaningful_words)
                feedback = (
                    f"A busca por '{term}' usou a frase completa como filtro único. "
                    f"Refaça usando cada palavra como AND separado com translate(): {and_filter}"
                )
                _log("supervisor", result=f"RETRY (frase multi-palavra exata) | {feedback[:80]}")
                return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # ── Fase 2: Análise de qualidade (sem retry — anota para o writer) ─────────

    # A. Contexto de extração ausente (validation error no extract)
    if not state.get("extracted_params") and rows:
        quality_notes.append(
            "Contexto de extração indisponível (erro de validação). "
            "Interprete os dados cuidadosamente."
        )
        _log("supervisor", result="AVISO: extracted_params=None com dados retornados")

    # B. Estoque: análise de saldos negativos e zerados
    if _ESTOQUE_RE.search(sql) and rows:
        qnt_idx = next((i for i, c in enumerate(col_names) if c == "QNT_ESTOQUE"), None)
        if qnt_idx is not None:
            try:
                valores = [float(r[qnt_idx]) for r in rows if r and r[qnt_idx] is not None]
                negativos = sum(1 for v in valores if v < 0)
                zerados   = sum(1 for v in valores if v == 0)
                positivos = sum(1 for v in valores if v > 0)
                total_val = sum(v for v in valores if v > 0)

                if negativos > 0:
                    quality_notes.append(
                        f"{negativos} produto(s) com saldo negativo (divergência de inventário — "
                        f"saídas não registradas). Não liste como estoque disponível."
                    )
                if zerados > 0 and positivos > 0:
                    quality_notes.append(
                        f"{zerados} produto(s) com saldo zerado além dos {positivos} com estoque."
                    )
                if positivos > 0:
                    quality_notes.append(f"Total de unidades com saldo positivo: {total_val:g}.")
            except (ValueError, TypeError):
                pass

    # B2. Datas ausentes → retry pedindo inclusão das colunas de período
    _NO_DATE_SKIP = {
        "certificado", "curva_abcd_lista", "curva_abcd_atencao",
        "curva_abcd_vs_mercado", "curva_abcd_sugestao_mix",
        "estoque_produto", "estoque_nivel",
        "pdv_versao", "pdv_config", "pdv_quantidade",
        "inconsistencias", "outro",
    }
    _DATE_COL_PREFIXES = (
        "DATA_INICIO", "DATA_FIM", "ANO_MES", "DIA", "MES_ATUAL",
        "PERIODO", "DATA_EMISSAO", "DATA_REFERENCIA",
    )
    if (
        rows and not sql_err and attempts < 3
        and metric not in _NO_DATE_SKIP
        and not any(
            any(col_n.startswith(p) for p in _DATE_COL_PREFIXES)
            for col_n in col_names
        )
    ):
        feedback = (
            "A query retornou dados mas sem colunas de período. "
            "Adicione obrigatoriamente as datas reais do filtro no SELECT conforme a regra 17: "
            "ex: date_sub(current_date(), 30) AS DATA_INICIO, date_sub(current_date(), 1) AS DATA_FIM. "
            "Para comparação de dois períodos use DATA_INICIO_ATUAL/DATA_FIM_ATUAL + DATA_INICIO_ANTERIOR/DATA_FIM_ANTERIOR."
        )
        _log("supervisor", result=f"RETRY (datas ausentes) | metric={metric}")
        return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # B3. Colunas de data presentes mas com valor NULL → retry
    # Acontece quando o SQL faz concat(i.ANO_MES, '-01') mas ANO_MES é DATE (não string),
    # produzindo '2026-05-01-01' inválido e retornando NULL para DATA_INICIO/DATA_FIM.
    _DATE_BOUND_COLS = {"DATA_INICIO", "DATA_FIM", "DATA_INICIO_ATUAL", "DATA_FIM_ATUAL"}
    _date_bound_present = [c for c in col_names if c in _DATE_BOUND_COLS]
    if _date_bound_present and rows and attempts < 3 and not sql_err:
        r0 = rows[0]
        null_date_cols = [
            c for c in _date_bound_present
            if r0[col_names.index(c)] is None
        ]
        if null_date_cols:
            feedback = (
                f"As colunas {null_date_cols} retornaram NULL. "
                "Causa provável: concat(i.ANO_MES, '-01') falha porque ANO_MES é DATE, não string. "
                "Corrija usando MIN(to_date(i.DATA_EMISSAO)) AS DATA_INICIO e "
                "MAX(to_date(i.DATA_EMISSAO)) AS DATA_FIM em vez de calcular a partir de ANO_MES."
            )
            _log("supervisor", result=f"RETRY (data cols NULL: {null_date_cols})")
            return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # C. Colunas inconsistentes com a métrica esperada
    _METRIC_EXPECTED_COLS = {
        "faturamento":      ("VALOR", "FAT", "FATURAMENTO"),
        "ticket_medio":     ("TICKET",),
        "gastos":           ("VALOR", "GASTO", "CUSTO"),
        "transacoes":       ("NOTAS", "TRANSAC", "CLIENTE"),
        "quantidade_vendida": ("QUANTIDADE", "QTD", "QNT"),
    }
    if metric in _METRIC_EXPECTED_COLS:
        expected = _METRIC_EXPECTED_COLS[metric]
        if not any(any(exp in col for exp in expected) for col in col_names):
            quality_notes.append(
                f"Atenção: colunas retornadas ({', '.join(col_names)}) "
                f"podem não corresponder à métrica '{metric}'. Responda com cautela."
            )

    # D. Estoque com resultado muito amplo (filtro muito genérico)
    if metric == "estoque_produto" and len(rows) > 30 and attempts < 3:
        product_filter = params_dict.get("product_filter") or ""
        if product_filter:
            feedback = (
                f"A query retornou {len(rows)} produtos para '{product_filter}' — filtro provavelmente muito amplo. "
                f"Torne o filtro mais específico: use termos mais longos ou combine com AND para reduzir resultados."
            )
            _log("supervisor", result=f"RETRY (estoque muito amplo: {len(rows)} linhas) | {feedback[:80]}")
            return {"supervisor_retry": True, "sql": None, "sql_error": feedback}

    # ── Fase 2.5: Validação semântica ────────────────────────────────────────
    # Só roda quando há dados, sem erro de SQL e ainda há tentativas disponíveis
    if rows and not sql_err and attempts < 3:
        relevant, reason = await _generator.check_relevance(
            question=question,
            metric=metric,
            columns=columns,
            rows=rows,
        )
        if not relevant:
            _log("supervisor", result=f"RETRY (semântico) | {reason[:80]}")
            return {"supervisor_retry": True, "sql": None, "sql_error": reason}

    # ── Fase 3: Sanitização ───────────────────────────────────────────────────

    # Trunca resultados excessivos para não estourar o contexto do writer
    # Produtos/estoque: máx 20. Dados analíticos (séries, comparações): máx 50.
    if _ESTOQUE_RE.search(sql):
        row_limit = 20
    elif grain in ("produto", "categoria"):
        row_limit = 20
    else:
        row_limit = 50
    if len(rows) > row_limit:
        _log("supervisor", result=f"TRUNCADO ({len(rows)} linhas → {row_limit})")
        return {
            "supervisor_retry": False,
            "rows": rows[:row_limit],
            "data_quality_notes": quality_notes,
            "answer": None,
        }

    # Certificado vencido → mensagem de renovação com link e suporte (pula para demo)
    cnpj_state = state.get("cnpj") or ""
    if metric == "certificado" and rows and not _is_demo_cnpj(cnpj_state):
        cert_val = str((rows[0] or [None])[0] or "").strip()
        if cert_val:
            try:
                cert_date = date.fromisoformat(cert_val[:10])
                if cert_date < date.today():
                    cnpj = state.get("cnpj") or ""
                    cnpj_digits = _cnpj_digits(cnpj)
                    cnpj_rev = cnpj_digits[::-1] if cnpj_digits else "00000000000000"
                    url_cert = (
                        "https://simtech.martins.com.br/"
                        f"SIMTECH.CONTRATO/Paginas/Questionarios/Index?1#{cnpj_rev}#1;2;15;16"
                    )
                    msg = (
                        f"Seu certificado digital está vencido desde {cert_date.strftime('%d/%m/%Y')}.\n\n"
                        f"Acesse o portal para renovar seu certificado digital:\n{url_cert}\n\n"
                        "Se tiver dúvidas, fale com o nosso suporte:\n"
                        "- WhatsApp: (34) 99912-7261\n"
                        "- Capitais e regiões metropolitanas: 3003-1266\n"
                        "- Outras regiões: 0800-729-5217"
                    )
                    _log("supervisor", result=f"certificado VENCIDO em {cert_date}")
                    return {"supervisor_retry": False, "direct_reply": msg}
            except (ValueError, TypeError, IndexError):
                pass

    if quality_notes:
        _log("supervisor", result=f"OK com notas: {quality_notes}")
    else:
        _log("supervisor", result="OK ✓")

    return {"supervisor_retry": False, "data_quality_notes": quality_notes}


_PERIOD_LABELS: dict[str, str] = {
    "ultimos_30_dias":  "nos últimos 30 dias",
    "este_mes":         "neste mês",
    "mes_passado":      "no mês passado",
    "semana_passada":   "na semana passada",
    "ontem":            "ontem",
    "hoje":             "hoje",
    "ano_passado":      "no ano passado",
    "dois_periodos":    "na comparação entre dois períodos",
    "periodo_custom":   "no período informado",
}


async def write_node(state: State) -> dict:
    """Formata a resposta final."""
    direct = (state.get("direct_reply") or "").strip()
    if direct:
        _log("write", source="direct_reply")
        return {"answer": direct}

    question = state.get("question") or ""
    columns  = state.get("columns") or []
    rows     = state.get("rows") or []

    params_dict    = state.get("extracted_params") or {}
    period_type    = params_dict.get("period_type") or ""
    period_detail  = params_dict.get("period_detail")
    period_detail_2 = params_dict.get("period_detail_2")
    metric         = params_dict.get("metric") or ""

    # Prioridade 1: se as linhas trazem DATA_INICIO_ATUAL/DATA_FIM_ATUAL + ANTERIOR,
    # monta o label com as datas reais retornadas pelo SQL.
    def _fmt_dt(v) -> str:
        s = str(v or "")
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
        return s

    cols_lower = [str(c).lower() for c in columns]
    period_label = None
    if (
        rows
        and "data_inicio_atual" in cols_lower and "data_fim_atual" in cols_lower
        and "data_inicio_anterior" in cols_lower and "data_fim_anterior" in cols_lower
    ):
        try:
            r0 = rows[0]
            ia = r0[cols_lower.index("data_inicio_atual")]
            fa = r0[cols_lower.index("data_fim_atual")]
            ip = r0[cols_lower.index("data_inicio_anterior")]
            fp = r0[cols_lower.index("data_fim_anterior")]
            period_label = (
                f"de {_fmt_dt(ia)} a {_fmt_dt(fa)} (vs {_fmt_dt(ip)} a {_fmt_dt(fp)})"
            )
        except Exception:
            period_label = None
    elif rows and "data_inicio" in cols_lower and "data_fim" in cols_lower:
        try:
            r0 = rows[0]
            di = r0[cols_lower.index("data_inicio")]
            df = r0[cols_lower.index("data_fim")]
            # só usa se ambos vieram com valor real — None/vazio faz a label virar "de  a "
            # o que é truthy e impede o fallback via period_detail abaixo
            if di and df:
                period_label = f"de {_fmt_dt(di)} a {_fmt_dt(df)}"
        except Exception:
            period_label = None

    # Fallback para diagnósticos e queries com dois períodos
    if not period_label and metric in ("diagnostico_positivo", "diagnostico_negativo"):
        today = date.today()
        d1_ini = today - timedelta(days=60)
        d1_fim = today - timedelta(days=31)
        d2_ini = today - timedelta(days=30)
        d2_fim = today - timedelta(days=1)
        period_label = (
            f"de {d2_ini.strftime('%d/%m/%Y')} a {d2_fim.strftime('%d/%m/%Y')} "
            f"(vs {d1_ini.strftime('%d/%m/%Y')} a {d1_fim.strftime('%d/%m/%Y')})"
        )
    elif not period_label and period_detail and period_detail_2:
        period_label = f"{period_detail} vs {period_detail_2}"
    elif not period_label and period_detail:
        period_label = period_detail
    elif not period_label:
        period_label = _PERIOD_LABELS.get(period_type)

    # Só mostra a nota de truncagem se o resultado atingiu exatamente o limite
    sql_note = state.get("sql_note")
    if sql_note:
        limit = (params_dict.get("limit") or 10)
        if len(rows) < limit:
            sql_note = None

    data_quality_notes = state.get("data_quality_notes") or []

    # Se a pergunta for muito curta (follow-up), enriquece com contexto dos params
    # para o writer conseguir aplicar o emoji e formato corretos
    if len(question.strip()) < 25 and params_dict:
        summary = params_dict.get("summary") or ""
        if summary:
            question = f"{question} [{summary}]"

    # Diagnostico sem resultados → pergunta o tema para gerar insight livre
    if metric in ("diagnostico_positivo", "diagnostico_negativo") and not rows:
        phone = state.get("phone") or ""
        if phone:
            from shared.session_store import set_pending_insight_theme
            set_pending_insight_theme(phone)
        tema_msg = (
            "📊 Não encontrei seções com destaque claro nos dados agora.\n\n"
            "Mas posso te dar um insight personalizado! *Sobre qual tema você gostaria?*\n\n"
            "Exemplos:\n"
            "• Faturamento e ticket médio\n"
            "• Produtos mais vendidos\n"
            "• Comparação com o mercado\n"
            "• Gestão de estoque\n"
            "• Tendências de vendas"
        )
        _log("write", source="diagnostico_vazio → pedindo tema")
        return {
            "answer":        tema_msg,
            "last_question": question,
            "last_answer":   tema_msg,
        }

    _log("write", source="LLM writer", rows=len(rows), cols=len(columns), period=period_label or "—")
    today = date.today().isoformat()
    answer = await _generator.write_answer(
        question=question, columns=columns, rows=rows,
        period_label=period_label, sql_note=sql_note, today=today,
        data_quality_notes=data_quality_notes,
    )
    _log("write", answer_len=len(answer), answer_preview=answer[:100])

    # Atualiza memória de longo prazo com a métrica e período usados
    cnpj_w  = state.get("cnpj") or ""
    params_w = state.get("extracted_params") or {}
    if cnpj_w and params_w.get("metric"):
        from shared.profile_store import update_profile
        update_profile(cnpj_w, params_w["metric"], params_w.get("period_type", ""))

    return {
        "answer":                answer,
        "last_question":         question,
        "last_answer":           answer,
        "last_extracted_params": state.get("extracted_params"),
    }


_NO_INSIGHT_METRICS = {
    "certificado", "estoque_produto", "estoque_nivel",
    "pdv_quantidade", "pdv_notas_processadas", "pdv_problemas",
    "pdv_versao", "pdv_config", "inconsistencias",
}


async def insight_node(state: State) -> dict:
    """Gera um insight de negócio acionável com base nos dados reais retornados."""
    # Insight automático em toda resposta está DESLIGADO por padrão (enchia o cliente).
    # O fluxo de insight sob demanda ("me dê um insight" → escolhe tema) continua ativo.
    # Para reativar o insight automático, defina AUTO_INSIGHT_ENABLED=1 no ambiente.
    if os.getenv("AUTO_INSIGHT_ENABLED", "0") != "1":
        return {"insight": None}

    rows    = state.get("rows") or []
    columns = state.get("columns") or []
    answer  = state.get("answer") or ""
    direct  = (state.get("direct_reply") or "").strip()
    params  = state.get("extracted_params") or {}
    metric  = params.get("metric", "")
    question = state.get("question") or ""

    # Não gera insight para: respostas diretas, sem dados ou métricas excluídas
    if direct or not rows or not answer or metric in _NO_INSIGHT_METRICS:
        return {"insight": None}

    # Serializa os dados reais para o LLM analisar
    col_names = [str(c) for c in columns]
    data_preview = "\n".join(
        "\t".join(str(v) for v in row)
        for row in rows[:10]
    )
    data_block = f"Colunas: {', '.join(col_names)}\nDados:\n{data_preview}"

    try:
        insight = await _generator.generate_insight(
            question=question,
            answer=f"{answer}\n\n[Dados estruturados]\n{data_block}",
            metric=metric,
            today=date.today().isoformat(),
        )
        _log("insight", preview=insight[:80] if insight else "vazio")
        return {"insight": insight or None}
    except Exception as e:
        _log("insight", result=f"ERRO: {e}")
        return {"insight": None}


async def insight_theme_node(state: State) -> dict:
    """Gera insight livre sobre o tema escolhido pelo usuário após diagnostico vazio."""
    theme   = state.get("question") or ""
    phone   = state.get("phone") or ""

    from shared.session_store import clear_pending_insight_theme
    if phone:
        clear_pending_insight_theme(phone)

    _log("insight_theme", theme=theme[:60])
    try:
        answer = await _generator.generate_themed_insight(theme=theme)
    except Exception as e:
        _log("insight_theme", result=f"ERRO: {e}")
        answer = "📊 Não consegui gerar o insight agora. Tente reformular o tema ou pergunte de outra forma."

    return {
        "answer":        answer,
        "last_question": theme,
        "last_answer":   answer,
    }


# ── Roteadores ─────────────────────────────────────────────────────────────────

def route_preprocess(state: State) -> str:
    if (state.get("direct_reply") or "").strip():
        return "write"
    intent = state.get("intent") or ""
    if intent == "insight_theme":
        return "insight_theme"
    if intent == "promocao":
        return "promocao"
    if intent.startswith("relatorio_") or intent.startswith("reposicao_"):
        return "hortifruti_report"   # cobre reposicao_, reposicao_mais_
    if intent.startswith("perdas_"):
        return "perdas_report"
    return "extract"


def route_sql_gen(state: State) -> str:
    if (state.get("direct_reply") or "").strip():
        return "write"     # resposta direta (ex: estoque não disponível)
    if state.get("sql"):
        return "execute"
    if (state.get("sql_attempts") or 0) < 3:
        return "sql_gen"   # retry
    return "write"         # desistiu


def route_supervisor(state: State) -> str:
    if state.get("supervisor_retry") and (state.get("sql_attempts") or 0) < 3:
        return "sql_gen"
    return "write"


# ── Build ──────────────────────────────────────────────────────────────────────

def route_cert_check(state: State) -> str:
    if (state.get("direct_reply") or "").strip():
        return "write"
    return "sql_gen"


def build_agent(checkpointer=None):
    from langgraph.checkpoint.memory import MemorySaver
    from langgraph.checkpoint.base import BaseCheckpointSaver
    # Guard: LangGraph API/Studio pode chamar build_agent(config_dict) — ignoramos
    if not isinstance(checkpointer, (BaseCheckpointSaver, type(None))):
        checkpointer = None
    if checkpointer is None:
        checkpointer = MemorySaver()

    g = StateGraph(State)

    g.add_node("preprocess",       preprocess_node)
    g.add_node("extract",          extract_node)
    g.add_node("cert_check",       cert_check_node)
    g.add_node("sql_gen",          sql_gen_node)
    g.add_node("execute",          execute_node)
    g.add_node("supervisor",       supervisor_node)
    g.add_node("write",            write_node)
    g.add_node("insight",          insight_node)
    g.add_node("hortifruti_report", hortifruti_report_node)
    g.add_node("perdas_report",    perdas_report_node)
    g.add_node("insight_theme",    insight_theme_node)
    g.add_node("promocao",         promocao_node)

    g.add_edge(START, "preprocess")
    g.add_edge("hortifruti_report", END)
    g.add_edge("perdas_report",    END)
    g.add_edge("insight_theme",    END)
    g.add_edge("promocao",         END)

    g.add_conditional_edges("preprocess", route_preprocess, {"write": "write", "extract": "extract", "hortifruti_report": "hortifruti_report", "perdas_report": "perdas_report", "insight_theme": "insight_theme", "promocao": "promocao"})
    g.add_edge("extract", "cert_check")
    g.add_conditional_edges("cert_check", route_cert_check, {"write": "write", "sql_gen": "sql_gen"})
    g.add_conditional_edges("sql_gen",    route_sql_gen,    {"execute": "execute", "sql_gen": "sql_gen", "write": "write"})
    g.add_edge("execute", "supervisor")
    g.add_conditional_edges("supervisor", route_supervisor, {"sql_gen": "sql_gen", "write": "write"})
    g.add_edge("write", "insight")
    g.add_edge("insight", END)

    return g.compile(checkpointer=checkpointer)
